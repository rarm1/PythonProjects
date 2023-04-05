from Resources import file_id_reader
import openpyxl as xl
import datetime
import sys

# TODO: We need to import the file reader. We then need to get the values, ideally keep things consistent.
#   This is going to look like doing things strictly by rows, which means I can iterate through rows
# TODO: Check all variables are in the list
global DOCUMENT, VAR_DICT

def obtain_file():
    all_files = file_id_reader.list_all_files(True)
    filename_user = file_id_reader.user_selected_file(all_files)
    filename = file_id_reader.file_name_generator(filename_user)[0]
    document = xl.load_workbook(filename, data_only=True)
    return document


def variable_definitions():
    global VAR_DICT
    var_sheet = DOCUMENT['Variables']
    variable_cell_list = [var_sheet.cell(row=a, column=1) for a in range(2, var_sheet.max_row+1) if
                  var_sheet.cell(row=a, column=1).value is not None]
    VAR_DICT = {a.value: a.offset(0, 1).value for a in variable_cell_list}


def harpoon():
    global DOCUMENT
    DOCUMENT = obtain_file()
    variable_definitions()
    written_row = 1
    output_sheet = DOCUMENT.create_sheet(title="OUTPUT")
    intro_paragraphs = paragraph_writer("Introduction")
    output_sheet.cell(row=written_row, column=1).value = "Introduction"
    written_row += 1
    for para in intro_paragraphs:
        output_sheet.cell(row=written_row, column=1).value = para
        written_row += 1
    asset_allocation_paragraphs = paragraph_writer("Asset Allocation")
    output_sheet.cell(row=written_row, column=1).value = "Asset Allocation"
    written_row += 1
    for para in asset_allocation_paragraphs:
        output_sheet.cell(row=written_row, column=1).value = para
        written_row += 1
    performance_paragraphs = paragraph_writer("Performance")
    output_sheet.cell(row=written_row, column=1).value = "Performance"
    written_row += 1
    for para in performance_paragraphs:
        output_sheet.cell(row=written_row, column=1).value = para
        written_row += 1
    costs_paragraphs = paragraph_writer("Ongoing Charges")
    output_sheet.cell(row=written_row, column=1).value = "Ongoing Charges"
    written_row += 1
    for para in costs_paragraphs:
        output_sheet.cell(row=written_row, column=1).value = para
        written_row += 1
    # TODO: This is where cost table needs to go.
    port_comp_sheet = DOCUMENT['Portfolio Comparison']
    margetts_ocf = 0.0
    client_ocf = 0.0
    client_name: str = str
    margetts_fund = []
    for i in range(1, 30):
        try:
            if 'Margetts Weighted OCF' in port_comp_sheet.cell(i, 7).value:
                margetts_ocf = port_comp_sheet.cell(i, 8).value
            if 'Weighted Average Client OCF' in port_comp_sheet.cell(i, 7).value:
                client_ocf = port_comp_sheet.cell(i, 8).value
            if 'Margetts Fund' in port_comp_sheet.cell(i, 7).value:
                margetts_fund.append(port_comp_sheet.cell(i, 8).value)
            if 'Margetts Share Class' in port_comp_sheet.cell(i, 7).value:
                margetts_fund.append(port_comp_sheet.cell(i, 8).value)
            if 'Client Name' in port_comp_sheet.cell(i, 7).value:
                client_name = port_comp_sheet.cell(i, 8).value
        except TypeError:
            # Because of the 'in' clause, this will error. Could be replaced with == when template is fully resolved.
            pass
    margetts_fund = " ".join(margetts_fund)
    output_sheet.cell(row=written_row, column=1).value = "Margetts Risk Rated " + margetts_fund
    output_sheet.cell(row=written_row, column=2).value = str(round(margetts_ocf*100, 1)) + "%"
    written_row += 1
    output_sheet.cell(row=written_row, column=1).value = client_name + " portfolio (weighted average underlying OCF)"
    output_sheet.cell(row=written_row, column=2).value = str(round(client_ocf * 100, 1)) + "%"
    DOCUMENT.save("Model Output.xlsx")


def higher_lower_calc(var):
    client_weight, margetts_weight = 0, 0
    if "margetts" in var:
        margetts_weight = VAR_DICT[var[:-13]]*100
        client_weight = VAR_DICT["client"+var[8:-13]]*100
    elif "client" in var:
        client_weight = VAR_DICT[var[:-13]]*100
        margetts_weight = VAR_DICT["margetts"+var[6:-13]]*100
    if "client" in var:
        if client_weight > margetts_weight and client_weight - margetts_weight > 2:
            return "higher"
        elif client_weight < margetts_weight and abs(client_weight - margetts_weight) > 2:
            return "smaller"
        else:
            return "similar"
    elif "margetts" in var:
        if margetts_weight > client_weight  and margetts_weight - client_weight > 2:
            return "higher"
        elif margetts_weight < client_weight and abs(margetts_weight- client_weight) > 2:
            return "smaller"
        else:
            return "similar"


def percieved_risk_calc(var):
    var = var[10:]
    client_weight, margetts_weight = 0, 0
    if "margetts" in var:
        margetts_weight = VAR_DICT[var]*100
        client_weight = VAR_DICT["client"+var[8]]*100
    elif "client" in var:
        client_weight = VAR_DICT[var]*100
        margetts_weight = VAR_DICT["margetts"+var[6:]]*100
    if "client" in var:
        if client_weight > margetts_weight and client_weight - margetts_weight > 2:
            return "lesser"
        elif client_weight < margetts_weight and abs(client_weight - margetts_weight) > 2:
            return "lesser"
        else:
            return "similar"
    elif "margetts" in var:
        if margetts_weight > client_weight  and margetts_weight - client_weight > 2:
            return "greater"
        elif margetts_weight < client_weight and abs(margetts_weight- client_weight) > 2:
            return "lesser"
        else:
            return "similar"


def annualised_calc():
    perf_comp = DOCUMENT['Performance Comparison']
    target_cells = []
    start = False
    for a in range(1, perf_comp.max_row+1):
        if start:
            target_cells.append([perf_comp.cell(a, 1).value, round(perf_comp.cell(a, 2).value,4),
                                 round(perf_comp.cell(a, 3).value, 4)])
        if perf_comp.cell(a, 1).value == "Annualised":
            start = True
    # 0 is year, 1 is margetts, 2 is client
    margetts_outperf_years = [str(a) for a, b, c in target_cells if b-c > 0.02]
    if len(margetts_outperf_years) >= 2:
        margetts_outperf_years[-1] = "and "+margetts_outperf_years[-1]
    margetts_outperf_years_str = ", ".join(margetts_outperf_years)
    client_outperf_years = [str(a) for a, b, c in target_cells if b-c < -0.02]
    if len(client_outperf_years) >= 2:
        client_outperf_years[-1] = "and "+client_outperf_years[-1]
    client_outperf_years_str = ", ".join(client_outperf_years)
    equal_years = [str(a) for a, b, c in target_cells if 0.02 > b - c > -0.02]
    if len(equal_years) >= 2:
        equal_years[-1] = "and "+equal_years[-1]
    equal_years_str = ", ".join(equal_years)
    to_return = ["The below chart shows discrete annual performance for the last 5 years."]
    if len(margetts_outperf_years) > 0:
        rr_outperformance = f"The Margetts Risk Rated {VAR_DICT['margetts_portfolio_identifier']}, " \
                        f"performed ahead of {VAR_DICT['client_name']}'s portfolio across {margetts_outperf_years_str}."
        to_return.append(rr_outperformance)
    if len(client_outperf_years) > 0:
        client_outperformance = f"{VAR_DICT['client_name']}'s portfolio outperformed the Margetts strategy in " \
                                f"{client_outperf_years_str}."
        to_return.append(client_outperformance)
    if len(equal_years) > 0:
        in_line = f"The two portfolios performed in line across {equal_years_str}."
        to_return.append(in_line)
    to_return = " ".join(to_return)
    return to_return


def performance_reporting(var):
    if "annualised_perf" in var:
        return annualised_calc()
    if "relative" not in var:
        try:
            test = str(round(VAR_DICT[var]*100, 1))
            return test
        except TypeError:
            print(f"var: {var}. Val: {VAR_DICT[var]}")

    margetts_perf = VAR_DICT[var[:4]+'performance_margetts']*100
    client_perf = VAR_DICT[var[:4]+'performance_client']*100
    margetts_outperformance = margetts_perf - client_perf

    if margetts_outperformance > 2:
        return 'ahead of'
    elif margetts_outperformance < -2:
        return 'behind'
    elif 2 >  margetts_outperformance > -2:
        return 'in line with'


def date_suffix(var):
    dictionary = {
        'st': [1, 21, 31],
        'nd': [2, 22],
        'rd': [3, 23]
    }
    for key in dictionary.keys():
        if var in dictionary[key]:
            return key
    else:
        return "th"


def client_risk_comparison():
    sheet = DOCUMENT['Portfolio Comparison']
    container = 0
    for col in range(1, sheet.max_column+1):
        container = sheet.cell(1, col)
        if container.value == "Risk Score":
            break
    try:
        margetts_risk = container.offset(1, 0).value
        client_risk = container.offset(2, 0).value
    except Exception as e:
        sys.exit("The template is not working as expected. Please ensure Risk Score in on Portfolio Comparison sheet "
                 f"in top row. {e}")
    if margetts_risk - client_risk > 1:
        return "takes less "
    elif margetts_risk - client_risk < -1:
        return "takes more "
    else:
        return "takes a similar amount of "
    pass


def paragraph_writer(sheetname):
    intro_sheet = DOCUMENT[sheetname]
    paragraph = []
    for a in range(2, intro_sheet.max_row+1):
        var = intro_sheet.cell(row=a, column=1).value
        if var in VAR_DICT.keys():
            # Predetermined (client name, portfolio name, etc)
            if isinstance(VAR_DICT[var], str):
                paragraph.append(VAR_DICT[var])
            # Perf date, date of request
            elif isinstance(VAR_DICT[var], datetime.datetime):
                day = VAR_DICT[var].strftime("%d")
                paragraph.append(str(VAR_DICT[var].strftime(f"%d{date_suffix(int(day))} %B %Y")))
            # Performance figure written straight to document no processing
            elif isinstance(VAR_DICT[var], float) and "performance" not in var:
                paragraph.append(str(round(VAR_DICT[var]*100, 1)))
            # Holding distribution comparison
            elif VAR_DICT[var] is None and "higher_lower" in var:
                paragraph.append(higher_lower_calc(var))
            # Percieved risk of portfolios
            elif VAR_DICT[var] is None and "percieved" in var:
                paragraph.append(percieved_risk_calc(var))
            # Performance comparisons
            elif "performance" in var:
                paragraph.append(performance_reporting(var))
            elif "client_portfolio_risk" in var:
                paragraph.append(client_risk_comparison())
        else:
            paragraph.append(var)
    paragraph = "".join(paragraph)
    paragraph = paragraph.split("new_line")
    return paragraph


if __name__ == '__main__':
    harpoon()
