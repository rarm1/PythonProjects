from datetime import datetime
from operator import itemgetter

import openpyxl
import os

global FILE, MOST_RECENT_FILENAME


def find_the_file():
    date_arr = []

    def date_convert(in_date):
        try:
            return datetime.strptime(in_date, '%d%m%y')
        except ValueError:
            return False

    for path, subdirs, files in os.walk("X:/Fund Management/Fund Management Dealing"):
        for name in files:
            try:
                date = date_convert(name[0:6])
                if date is not False:
                    date_arr.append([date, name])
            except ValueError:
                pass

    sorted_arr = sorted(date_arr, key=itemgetter(0), reverse=True)
    latest_date = sorted_arr[0][1][0:6]

    files_from_date = [file for file in sorted_arr[0:10] if file[1][0:6] == latest_date]

    def sort_by_time(arr):
        return arr[1][7:11]

    sorted_arr = sorted(files_from_date, key=sort_by_time, reverse=True)[0:2]
    try:
        if sorted_arr[0][1][0:11] == sorted_arr[1][1][0:11]:
            if sorted_arr[0][1][-9:-5] > sorted_arr[1][1][-9:-5]:
                return sorted_arr[0][1]
            else:
                return sorted_arr[1][1]
        else:
            return sorted_arr[0][1]
    except IndexError:
        return sorted_arr[0][1]


def check_cashflows():
    document = FILE
    overweight_share_class = []
    for row in range(1, document.max_row):
        record_type = document.cell(row, 1)
        if str(record_type.value) == '2':
            try:
                if abs(record_type.offset(0, 13).value / record_type.offset(0, 4).value) > 0.005:
                    overweight_share_class.append(document.cell(row, 3))
            except ZeroDivisionError:
                pass
    return overweight_share_class


def portfolio_weighting_calculator(overweight_classes):
    overweight_portfolios = []
    for shareclass in overweight_classes:
        total_units = 0
        unit_movement = 0
        cash_flow = 0
        for row in range(1, FILE.max_row):
            record_type = FILE.cell(row, 1)
            if str(record_type.value) == '2':
                if shareclass.offset(0, -1).value[:2] == FILE.cell(row, 2).value[0:2] and \
                        shareclass.value[:9] == FILE.cell(row, 3).value[0:9]:
                    total_units += record_type.offset(0, 4).value
                    unit_movement += record_type.offset(0, 13).value
                    cash_flow += record_type.offset(0, 18).value
        else:
            shareclass_weighting = ((unit_movement / total_units) * 100)
            if abs(shareclass_weighting) > 0.5:
                overweight_portfolios.append([shareclass, shareclass_weighting, cash_flow])
    else:
        return overweight_portfolios


def main():
    global FILE
    filename = find_the_file()
    print(f'File read: {filename}')
    FILE = openpyxl.load_workbook("X:/Fund Management/Fund Management Dealing/" + filename, data_only=True)
    FILE = FILE.active
    overweight_classes = check_cashflows()
    portfolios_to_check = portfolio_weighting_calculator(overweight_classes)
    if len(portfolios_to_check) == 0:
        print("No portfolios of concern.")
    else:
        for portfolio, weighting, cashflow in portfolios_to_check:
            print(f"Portfolio: {portfolio.value}, movement of {round(weighting, 4)}%, £{cashflow}")


if __name__ == '__main__':
    main()
