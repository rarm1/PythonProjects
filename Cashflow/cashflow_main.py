import os
from datetime import datetime
from operator import itemgetter

import openpyxl

global FILE, MOST_RECENT_FILENAME

PROVIDENCE = ['PAC', 'PIN', 'PRO', 'PSA', 'PSI']
SELECT = ['SEL', 'SLA', 'SSA']
VENTURE = ['VEN', 'VNA', 'VSA']
INTERNATIONAL = ['IAS', 'INS', 'INT']


def find_the_file():
	"""

	:return:
	"""
	date_arr = []
	
	def date_convert(in_date):
		"""

		:param in_date:
		:return:
		"""
		try:
			return datetime.strptime(in_date, '%d%m%y')
		except ValueError:
			return False
	
	for path, subdirs, files in os.walk("X:/Fund Management/Fund Management Dealing"):
		for name in files:
			try:
				date = date_convert(name[0:6])
				if date is not False:
					date_arr.append([date, name])
			except ValueError:
				pass
	
	sorted_arr = sorted(date_arr, key=itemgetter(0), reverse=True)
	latest_date = sorted_arr[0][1][0:6]
	
	files_from_date = [file for file in sorted_arr[0:10] if file[1][0:6] == latest_date]
	
	def sort_by_time(arr):
		"""

		:param arr:
		:return:
		"""
		return arr[1][7:11]
	
	sorted_arr = sorted(files_from_date, key=sort_by_time, reverse=True)[0:2]
	try:
		if sorted_arr[0][1][0:11] == sorted_arr[1][1][0:11]:
			if sorted_arr[0][1][-9:-5] > sorted_arr[1][1][-9:-5]:
				return sorted_arr[0][1]
			else:
				return sorted_arr[1][1]
		else:
			return sorted_arr[0][1]
	except IndexError:
		return sorted_arr[0][1]


def check_cashflows():
	"""

	:return:
	"""
	document = FILE
	overweight_share_class = []
	for row in range(1, document.max_row):
		record_type = document.cell(row, 1)
		if str(record_type.value) == '2':
			try:
				if abs(record_type.offset(0, 13).value / record_type.offset(0, 4).value) > 0.005:
					overweight_share_class.append(document.cell(row, 3))
			except ZeroDivisionError:
				pass
	return overweight_share_class


def convert_back_to_cells(asset_code_arr):
	"""

	:param asset_code_arr:
	:return:
	"""
	# So I think in this function I basically want to convert from the asset code array (not RR)
	for i, asset_code in enumerate(asset_code_arr):
		for row in range(1, FILE.max_row):
			record_type = FILE.cell(row, 1)
			asset_code_cell = FILE.cell(row, 2)
			if str(record_type.value) == '2' and asset_code_cell.value == asset_code:
				asset_code_arr[i] = asset_code_cell
	return asset_code_arr


def fund_difference_finder(given_range, range_name):
	"""

	:param given_range:
	:param range_name:
	:return:
	"""
	# TODO: For each fund in the given range, find the difference. i.e. for fund in given range, for each fund in file
	total_units = 0
	unit_movement = 0
	cash_flow = 0
	overweight_portfolios = []
	for fund in given_range:
		for row in range(1, FILE.max_row):
			record_type = FILE.cell(row, 1)
			asset_code = FILE.cell(row, 2)
			if str(record_type.value) == '2' and asset_code.value == fund:
				total_units += record_type.offset(0, 4).value
				unit_movement += record_type.offset(0, 13).value
				cash_flow += record_type.offset(0, 18).value
		else:
			shareclass_weighting = ((unit_movement / total_units) * 100)
			if abs(shareclass_weighting) > 0.5:
				overweight_portfolios.append([range_name, shareclass_weighting, cash_flow])
	return overweight_portfolios


def limited_shareclass_info(arr):
	"""

	:param arr:
	:return:
	"""
	asset_code_arr = [i.offset(0, -1).value for i in arr]
	rr_diffs = []
	for asset_code in asset_code_arr:
		if asset_code in PROVIDENCE:
			print("Prov")
		
		elif asset_code in SELECT:
			# for _ in arr:
			#     if not _.offset(0, -1).value == asset_code and _.offset(0, -1).value in SELECT:
			# TODO: I think the way to get around this is to do something like find every asset
			select_diff = fund_difference_finder(SELECT, "Select")
			rr_diffs.append(select_diff)
			for asset_code_list in SELECT:
				try:
					asset_code_arr.remove(asset_code_list)
				except ValueError:
					pass
		
		elif asset_code in VENTURE:
			print("Venture ")
		
		elif asset_code in INTERNATIONAL:
			print("International")
	return_arr = convert_back_to_cells(asset_code_arr)
	# Return array of cells as before without the RR range. And return RR Portfolios if over 0.5%
	return return_arr, rr_diffs


def portfolio_weighting_calculator(overweight_classes):
	"""

	:param overweight_classes:
	:return:
	"""
	overweight_portfolios = []
	# TODO: Would it make more sense to append to overweight portfolios here, and reduce the size of the arr here too?
	# overweight_classes, risk_rated_portfolios = limited_shareclass_info(overweight_classes)
	# print(risk_rated_portfolios)
	for shareclass in overweight_classes:
		total_units = 0
		unit_movement = 0
		cash_flow = 0
		for row in range(1, FILE.max_row):
			record_type = FILE.cell(row, 1)
			if str(record_type.value) == '2':
				if shareclass.offset(0, -1).value[:2] == FILE.cell(row, 2).value[0:2] and \
						shareclass.value[:9] == FILE.cell(row, 3).value[0:9]:
					total_units += record_type.offset(0, 4).value
					unit_movement += record_type.offset(0, 13).value
					cash_flow += record_type.offset(0, 18).value
		else:
			shareclass_weighting = ((unit_movement / total_units) * 100)
			if abs(shareclass_weighting) > 0.5:
				overweight_portfolios.append([shareclass, shareclass_weighting, cash_flow])
	else:
		return overweight_portfolios


def main():
	"""

	"""
	global FILE
	filename = find_the_file()
	print(f'File read: {filename}')
	FILE = openpyxl.load_workbook("X:/Fund Management/Fund Management Dealing/" + filename, data_only=True)
	FILE = FILE.active
	# FILE = openpyxl.load_workbook("Intra Day Cash Flow Report - 270522 0830VP.XLSX", data_only=True)
	# FILE = FILE.active
	overweight_classes = check_cashflows()
	portfolios_to_check = portfolio_weighting_calculator(overweight_classes)
	if len(portfolios_to_check) == 0:
		print("No portfolios of concern.")
	else:
		for portfolio, weighting, cashflow in portfolios_to_check:
			print(f"Portfolio: {portfolio.value}, movement of {round(weighting, 4)}%, £{cashflow}")


if __name__ == '__main__':
	main()
