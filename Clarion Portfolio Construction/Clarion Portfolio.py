import openpyxl as xl

targets_file = xl.load_workbook('Copy of Clarion Targets Updated 13 07 2022 v2.xlsx', data_only=True)
prudence_targets = targets_file['Prudence']
navigator_targets = targets_file['Navigator']
meridian_targets = targets_file['Meridian']
explorer_targets = targets_file['Explorer']
prudence_rebalancer = xl.load_workbook('Prudence Portfolio.xlsx', data_only=True).active
navigator_rebalancer = xl.load_workbook('Navigator Portfolio.xlsx', data_only=True).active
prudence_max_rows = 37
for rebalancer_row in range(1, prudence_max_rows):
    isin_found = False
    rebalancer_isin = prudence_rebalancer.cell(row=rebalancer_row, column=3).value
    for targets_row in range(8, navigator_targets.max_row):
        targets_isin = navigator_targets.cell(row=targets_row, column=2).value
        if rebalancer_isin is not None and rebalancer_isin == targets_isin:
            isin_found = True
    else:
        if not isin_found and rebalancer_isin is not None:
            print(f"Can't find this: {rebalancer_isin} which is on {rebalancer_row}. Which is {prudence_rebalancer.cell(row=rebalancer_row, column=5).value}")



