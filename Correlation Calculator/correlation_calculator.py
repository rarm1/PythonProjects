# Library imports
import math
from operator import itemgetter

import numpy
from scipy.stats import pearsonr
import openpyxl as xl
import pandas as pd

# Local imports
from Resources import file_id_reader

# Global variable declaration
global ALL_FUNDS, ALL_RETURNS, USER_DEFINED_FUND, READER_SHEET, USER_DEF_FUND_RETS, NON_BROKER_FILENAME, HOLDING_PERIODS


# This function creates multiple dataframes based on y1, y3 and y5 hence the dataframe names.
def correlation_writer():
    # Generates correlation using the multiloop function.
    correls = multi_loop()
    # Variable initialisation.
    y1, y3, y5 = [], [], []
    # Loop through all funds in correlation.
    for fund in correls:
        y1.append([fund[0], fund[1][0][0], fund[1][0][1]])
        y3.append([fund[0], fund[2][0][0], fund[1][0][1]])
        y5.append([fund[0], fund[3][0][0], fund[1][0][1]])

        # if not ((numpy.isnan(fund[1][0][0]))):
        #     y1.append([fund[0], fund[1][0][0], fund[1][0][1]])
        # else:
        #     y1.append([fund[0], -2, -2])
        #
        # if not numpy.isnan(fund[2][0][0]):
        #     y3.append([fund[0], fund[2][0][0], fund[1][0][1]])
        # else:
        #     y3.append([fund[0], -2, -2])
        #
        # if not numpy.isnan(fund[3][0][0]):
        #     y5.append([fund[0], fund[3][0][0], fund[1][0][1]])
        # else:
        #     y5.append([fund[0], -2, -2])

    # Sort these arrays by the correlation. This can be changed to itemgetter 2 to order by accuracy estimate.
    y1 = sorted(y1, key=itemgetter(1), reverse=True)
    y3 = sorted(y3, key=itemgetter(1), reverse=True)
    y5 = sorted(y5, key=itemgetter(1), reverse=True)
    # Generate dataframes to be written to Excel output.
    df1 = pd.DataFrame(data=y1, columns=["Investment Name", "Correlation", "Accuracy Estimate"])
    df3 = pd.DataFrame(data=y3, columns=["Investment", "Correlation", "Accuracy Estimate"])
    df5 = pd.DataFrame(data=y5, columns=["Investment", "Correlation", "Accuracy Estimate"])
    df1 = df1.dropna()
    df3 = df3.dropna()
    df5 = df5.dropna()
    return df1, df3, df5


# Creates and writes to Excel document
def excel_printer(to_write, ):
    # Create final filename for file output.
    filename = NON_BROKER_FILENAME[:-5] + " Finished.xlsx"
    # Write dataframes to Excel.
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Writes One, Three and Five year correlations to Excel.
        pd.DataFrame(to_write[0]).to_excel(writer, index=False, sheet_name="Year One Correlation")
        pd.DataFrame(to_write[1]).to_excel(writer, index=False, sheet_name="Year Three Correlation")
        pd.DataFrame(to_write[2]).to_excel(writer, index=False, sheet_name="Year Five Correlation")
    print("Correlation Write Successful")


# Finds the ISIN given by the user to correlate other funds to.
def find_objective_isin(requested_isin):
    # Search through the given Excel document to find the user given fund.
    for x in range(1, READER_SHEET.max_row):
        if READER_SHEET.cell(x, 2).value == requested_isin:
            return READER_SHEET.cell(x, 1), True
    print("Incorrect ISIN / not found.")
    return 0, False


# Finds the company the user gave, allows to ensure that the user gives an ISIN contained with the document.
def isin_checker():
    objective_isin = False
    user_def_fund = []
    while not objective_isin:
        user_input_isin = input("Enter ISIN: ")
        user_def_fund = find_objective_isin(user_input_isin)
        objective_isin = user_def_fund[1]
    return user_def_fund[0]


def generate_fund_arr(find_returns_for_this_fund):
    """
    Returns for funds which are not requested by the user.
    :param find_returns_for_this_fund: Takes fund from loop.
    :return: Returns array of returns.
    """
    global READER_SHEET
    fund_to_compare_arr = []
    if USER_DEFINED_FUND != find_returns_for_this_fund:
        for column in file_id_reader.returns_column():
            cell = READER_SHEET.cell(row=find_returns_for_this_fund.row, column=column.column)
            if cell.value is None:
                cell.value = 0
            else:
                cell.value /= 100
            fund_to_compare_arr.append(cell.value)
    return fund_to_compare_arr


# This is the function that does all the heavy lifting.
def generate_pearson(fund_in):
    """
    This function processes information to pass to the pearsonr calculator from the 'stats' library.
    :param fund_in: Accepts one funds returns
    :return: Returns the correlation for 1, 3 and 5 year returns.
    """
    # Variable initialisation.
    global USER_DEF_FUND_RETS
    y1, y3, y5 = [], [], []
    # Loops through years. 3 time periods are used.
    for idx, year in enumerate([1, 3, 5]):
        # This uses the 'stats' library to generate the correlation between funds.
        pearson_correl = pearsonr(USER_DEF_FUND_RETS[:year * 52], fund_in[:year * 52])
        # check_if_na = any(map(lambda element: math.isnan(element), pearson_correl))
        # test_replace = lambda x: math.isnan(x), pearson_correl
        # for i, x in enumerate(pearson_correl):
        #     if math.isnan(x):
        #         pearson_correl[][i] = 0
            # if not check_if_na:

        # for x in pearson_correl:
        #     if math.isnan(x):
        #         print(x)
        #         x = 0
        if idx == 0:
            y1.append(pearson_correl)
        elif idx == 1:
            y3.append(pearson_correl)
        elif idx == 2:
            y5.append(pearson_correl)

    return y1, y3, y5


# This loop runs through all funds and then offloads the hardwork to other functions.
def multi_loop():
    # Variable initialisation
    to_return = []
    # This finds all funds and runs through them in an iterative manner.
    for fund in file_id_reader.find_all_funds():
        if fund.row != USER_DEFINED_FUND.row:
            # This generates the full array of returns for each fund.
            fund_arr = generate_fund_arr(fund)
            pearson_correl = generate_pearson(fund_arr)
            y1 = (pearson_correl[0])
            y3 = (pearson_correl[1])
            y5 = (pearson_correl[2])
            fund_name = fund.offset(0, -1).value
            to_return.append([fund_name, y1, y3, y5])
    return to_return


# This is the key function, it could be integrated into main, but splitting it allows for easier digestion and readability.
def execution_init():
    # Variable initiation
    global USER_DEFINED_FUND, USER_DEF_FUND_RETS
    # Get the isin to correlation to, from the user.
    USER_DEFINED_FUND = isin_checker()
    # Generate the returns from for the fund specified by the user.
    USER_DEF_FUND_RETS = generate_fund_arr(USER_DEFINED_FUND.offset(0, 1))
    to_write = correlation_writer()
    excel_printer(to_write)


# Main function that is functionally similar to other main functions. Get the file from the user, then offload the
# Rest of the functionality to other functions. #
def main():
    # Variable initialisation
    global ALL_FUNDS, ALL_RETURNS, READER_SHEET, NON_BROKER_FILENAME
    # all_files = file_id_reader.list_files_by_type(True)
    # filename_user = file_id_reader.select_file_by_user(all_files)
    filenames = file_id_reader.file_name_generator('Correlation Data One.xlsx')
    for filename in filenames:
        FILENAME = filename
        NON_BROKER_FILENAME = FILENAME[:-5]+'.xlsx'
        reader_doc = xl.load_workbook(filename)
        reader_sheet = reader_doc.active
        READER_SHEET = reader_sheet
        file_id_reader.READER_SHEET = reader_sheet
        ALL_FUNDS = file_id_reader.find_all_funds()
        ALL_RETURNS = file_id_reader.returns_column()
        execution_init()


# In line with PEP8 standards.
if __name__ == '__main__':
    main()
