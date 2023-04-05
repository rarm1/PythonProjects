# -*- coding: utf-8 -*-
"""
Created on Wed Aug  3 09:54:21 2022

@author: Richard Armstrong
"""
# TODO: This reads in as openpyxl. Which is great for usability but shit for speed. Can this be cleaned up?
# Break down into 8 classifications.
# Imports
from operator import itemgetter
import openpyxl as xl
import statistics as stats
import numpy as np
import pandas as pd

from Resources import file_id_reader

# Variable declaration: as per PEP 8 global variables are capitalised.
global READER_SHEET, ALL_FUNDS, ALL_RETURNS, MODIFIED_FILENAME, FILENAME


# This creates a dictionary of returns against each fund.
def dictionary_creation():
    funds_dict = {}
    for fund in ALL_FUNDS:
        if fund.offset(0, 1).value is None or fund.offset(0, 2).value is None:
            continue
        fund_returns = []
        for week in ALL_RETURNS:
            cell_val = READER_SHEET.cell(row=fund.row, column=week.column).value
            if cell_val is not None:
                fund_returns.append(READER_SHEET.cell(row=fund.row, column=week.column))
        else:
            funds_dict[fund] = fund_returns
    else:
        return funds_dict


def process_weekly_returns(fund_returns):
    """
    Processes all returns into a usable format including calculating total returns
    :param fund_returns: Takes a list of returns
    :return: all_mr, three_year_returns, five_year_returns
    """
    # Variable declaration
    count_y5_per_fund = 0
    w1, w2, w3, w4, all_mr, three_year_returns, five_year_returns = 100, 100, 100, 100, [], None, None
    # Uses a generator style function to calculate all fund returns from array
    return_calculations = [(a.value / 100) + 1 for a in fund_returns]
    for idx, weekly in enumerate(return_calculations):
        # This calculates returns by referencing the cell prior, hence why it ensures there is a cell suitably far behind.
        w4 = (w4 * return_calculations[idx])
        if idx > 0:
            w3 = (w3 * return_calculations[idx - 1])
        if idx > 1:
            w2 = (w2 * return_calculations[idx - 2])
        if idx > 2:
            w1 *= return_calculations[idx - 3]
            all_mr.append(monthly_returns(w1, w2, w3, w4))
        if idx == 156:  # 52 * 3
            three_year_returns = w4
        elif idx > 255:  # 52 *5
            five_year_returns = w4
    else:
        return all_mr, three_year_returns, five_year_returns


# This function calculates min and max equity as well as the difference at the end of the month, from these high/low points.
def monthly_returns(w1v, w2v, w3v, w4v):
    # The end point of the month.
    e = w4v
    # Peak equity
    pe = max([w1v, w2v, w3v, w4v])
    # Minimum equity
    me = min([w1v, w2v, w3v, w4v])
    # Peak equity minus final equity value over the peak equity.
    mrpp = (pe - e) / pe
    # End equity minus minimum equity over final equity.
    mrsl = (e - me) / e
    # The highest value of the MR PP and the MRS L
    mr = max([mrpp, mrsl])
    return mr


# Returns the AMR for the third and fifth year values.
def amr_calculator(amr, year_3_value, year_5_value):
    # print(amr[0:5], year_3_value, year_5_value)
    # This ensures the value is recorded as 0 but elimates an error.
    if len(amr) <= 10:
        return 0, 0, 0
    else:
        # This finds the average of the AMR's that're passed to the array.
        completed_amr = stats.mean(amr)
    if year_3_value is None:
        year_3_value = 0
    if year_5_value is None:
        year_5_value = 0
    third = (year_3_value - 100) / (completed_amr * 100)
    fifth = (year_5_value - 100) / (completed_amr * 100)
    return third, fifth


# This formats the output to be passed to the Excel writer.
def output_generator(company_being_read, amr, amount_returns):
    # print(company_being_read, amr, amount_returns)
    company_name = company_being_read.offset(0, -1).value
    isin = company_being_read.value
    average_cap = company_being_read.offset(0, 1).value
    g_v_s = company_being_read.offset(0, 2).value
    year_three = amr[0]
    year_five = amr[1]
    return [company_name, isin, average_cap, g_v_s, year_three, year_five, amount_returns]


# This runs through each fund and returns the structured output.
def generator(fund_dictionary):
    to_return = []
    keys_to_remove = []
    for key in fund_dictionary.keys():
        amr, three_y, five_y = process_weekly_returns(fund_dictionary[key])
        if five_y is None:
            keys_to_remove.append(key)
            continue
        amr_returns = amr_calculator(amr, three_y, five_y)
        structured_output = output_generator(key, amr_returns, len(fund_dictionary[key]))
        to_return.append(structured_output)
    for key in keys_to_remove:
        fund_dictionary.pop(key)
    return to_return


# There are 3 groups that the value growth scores are to be broken into.
def find_value_growth_score_sheets(fund_array):
    to_return = []
    for fund in fund_array:
        if fund[3] is None:
            continue
        try:
            float(fund[3])
        except ValueError as e:
            # print(e)
            continue
        to_return.append(fund[3])
    return np.quantile(to_return, q=np.arange(0, 1, 1/3))[1:]


# Groups all funds by their growth/value scores.
def sort_growth_value_scores(funds, boundaries):
    value, balanced, growth = [], [], []
    for fund in funds:
        if fund[3] is None:
            continue
        try:
            float(fund[3])
        except ValueError:
            continue
        if fund[3] < boundaries[0]:
            value.append(fund)
        elif boundaries[0] < fund[3] < boundaries[1]:
            balanced.append(fund)
        else:
            growth.append(fund)
    return [value, balanced, growth]


def high_cap_low_cap(grow_bal_val_sorted):
    value_low_cap, value_mid, value_high_cap = [], [], []
    balanced_low_cap, balanced_mid, balanced_high_cap = [], [], []
    growth_low_cap, growth_mid, growth_high_cap,  = [], [], []
    for idx, v_g_s in enumerate(grow_bal_val_sorted):
        for fund in v_g_s:
            if fund[2] is None:
                fund[2] = 0
            if fund[2] > 6000:
                if idx == 0:
                    value_high_cap.append(fund)
                elif idx == 1:
                    balanced_high_cap.append(fund)
                elif idx == 2:
                    growth_high_cap.append(fund)
            elif 2000 > fund[2]:
                if idx == 0:
                    value_low_cap.append(fund)
                elif idx == 1:
                    balanced_low_cap.append(fund)
                elif idx == 2:
                    growth_low_cap.append(fund)
            else:
                if idx == 0:
                    value_mid.append(fund)
                elif idx == 1:
                    balanced_mid.append(fund)
                elif idx == 2:
                    growth_mid.append(fund)
    return [value_low_cap, value_mid, value_high_cap,
            balanced_low_cap, balanced_mid, balanced_high_cap,
            growth_low_cap, growth_mid, growth_high_cap]


def filter_the_best(fully_sorted_companies: [list], companies_per_sector) -> list:
    """
    :type fully_sorted_companies: list of lists of lists
    :rtype list of lists
    """
    # filtered_companies = [classification[:5] if idx % 2 == 1 else classification[:10] for idx, classification in enumerate(fully_sorted_companies, start=1)]
    # company_names = [[i[0] for i in classification] for classification in filtered_companies]

    filtered_companies = [classification[:companies_per_sector] for idx, classification in enumerate(fully_sorted_companies,
                                                                                        start=1)]
    company_names = [[i[0] for i in classification] for classification in filtered_companies]
    return company_names


def return_all_companies(fully_sorted_companies):
    list_companies = [classification for idx, classification in enumerate(fully_sorted_companies, start=1)]
    all_company_names = [[i[0] for i in classification] for classification in list_companies]
    return all_company_names


def modify_reader_sheet(arr_in):
    df = pd.read_excel(FILENAME, skiprows=8, index_col=0, header=0).T
    df = df.dropna(subset=["Value-Growth Score (Long)"], inplace=False, axis=1)
    sector = df['IA UK All Companies']
    desired_list = [x for sublist in arr_in for x in sublist]
    for catagorisation in df.columns:
        if catagorisation not in desired_list:
            df.pop(catagorisation)
    return df, sector


def strip_main():
    # So I need to think about whether dataframes are the best way to do this calculation going forward.
    #   that could be the first thing that I try and htink about improving if the program is dead slow.
    global READER_SHEET, ALL_FUNDS, ALL_RETURNS, FILENAME
    all_files = file_id_reader.list_all_files(False)
    FILENAME = file_id_reader.user_selected_file(all_files, enable_all=True, preselected='Smaller Companies Data.xlsx')
    reader_doc = xl.load_workbook(FILENAME)
    READER_SHEET = reader_doc.active
    file_id_reader.READER_SHEET = READER_SHEET
    ALL_FUNDS = file_id_reader.find_all_funds()
    ALL_RETURNS = file_id_reader.returns_column()
    funds_dict = dictionary_creation()
    output_ready_for_processing = generator(funds_dict)
    deciled_output = sorted(output_ready_for_processing, key=itemgetter(5), reverse=True)
    grow_bal_val_boundaries = find_value_growth_score_sheets(deciled_output)
    grow_bal_val_sorted = sort_growth_value_scores(deciled_output, grow_bal_val_boundaries)
    fully_sorted_companies = high_cap_low_cap(grow_bal_val_sorted)
    best_filtered_companies = filter_the_best(fully_sorted_companies, companies_per_sector=10)  # Filter the top XXX
    # amount of companies by their RRR Ranking
    best_filtered_companies = return_all_companies(fully_sorted_companies)  # Takes all companies to provide a more broad scope.
    # print(len(best_filtered_companies))
    # print(len(all_companies))
    finished_df, sector = modify_reader_sheet(best_filtered_companies)
    if __name__ == '__main__':
        finished_df.T.to_csv("out.csv")
    # finished_df, sector = modify_reader_sheet(fully_sorted_companies)
    # if __name__ == '__main__':
    #     finished_df.T.to_csv("in.csv")
    return finished_df, sector


if __name__ == '__main__':
    strip_main()
