# Imports
import statistics as stats
from operator import itemgetter

import numpy as np
import openpyxl as xl

from Resources import file_id_reader

# TODO: Refactor to be object oriented. Unnecessarily complicated at the moment.
global READER_SHEET, ALL_FUNDS, ALL_RETURNS, MODIFIED_FILENAME


# This creates a dictionary of returns against each fund.
def dictionary_creation():
    funds_dict = {}
    for fund in ALL_FUNDS:
        fund_returns = []
        for week in ALL_RETURNS:
            # if READER_SHEET.cell(row=fund.row, column=week.column).value is not None:
            fund_returns.append(READER_SHEET.cell(row=fund.row, column=week.column))
        else:
            funds_dict[fund] = fund_returns
    else:
        return funds_dict


def process_weekly_returns(fund_returns):
    """
    Processes all returns into a usable format including calculating total returns
    :param fund_returns: Takes a list of returns
    :return: all_mr, three_year_returns, five_year_returns
    """
    # Variable declaration
    w1, w2, w3, w4, all_mr, three_year_returns, five_year_returns = 100, 100, 100, 100, [], None, None
    # Uses a generator style function to calculate all fund returns from array
    return_calculations = [(a.value / 100) + 1 for a in fund_returns]
    for idx, weekly in enumerate(return_calculations):
        # if weekly is not None:
        #     started = True
        # if started is False and weekly is None:
        #     continue
        # if started and weekly is None:
        #     print("Confirm this is returning none correctly", key)
        # This allows the user to check that the value is none.
        if weekly is None:
            print(weekly, idx)
        # This calculates returns by referencing the cell prior, hence why it ensures there is a cell suitably far behind.
        w4 = (w4 * return_calculations[idx])
        if idx > 0:
            w3 = (w3 * return_calculations[idx - 1])
        if idx > 1:
            w2 = (w2 * return_calculations[idx - 2])
        if idx > 2:
            w1 *= return_calculations[idx - 3]
            all_mr.append(monthly_returns(w1, w2, w3, w4))
        if idx == (52 * 3):
            three_year_returns = w4
        elif idx == (52 * 5):
            five_year_returns = w4
        # if started and weekly is None:
        #     print("Confirm this is returning none correctly", key)
    else:
        return all_mr, three_year_returns, five_year_returns


# This function calculates min and max equity as well as the difference at the end of the month, from these high/low points.
def monthly_returns(w1v, w2v, w3v, w4v):
    # The end point of the month.
    e = w4v
    # Peak equity
    pe = max([w1v, w2v, w3v, w4v])
    # Minimum equity
    me = min([w1v, w2v, w3v, w4v])
    # Peak equity minus final equity value over the peak equity.
    mrpp = (pe - e) / pe
    # End equity minus minimum equity over final equity.
    mrsl = (e - me) / e
    # The highest value of the MR PP and the MRS L
    mr = max([mrpp, mrsl])
    return mr


# Returns the AMR for the third and fifth year values.
def amr_calculator(amr, year_3_value, year_5_value):
    # This ensures the value is recorded as 0 but elimates an error.
    if not (len(amr)) > 10:
        return 0, 0, 0
    else:
        # This finds the average of the AMR's that're passed to the array.
        completed_amr = stats.mean(amr)
    if year_3_value is None:
        year_3_value = 0
    if year_5_value is None:
        year_5_value = 0
    third = (year_3_value - 100) / (completed_amr * 100)
    fifth = (year_5_value - 100) / (completed_amr * 100)
    return third, fifth


# This formats the output to be passed to the Excel writer.
def output_generator(company_being_read, amr, amount_returns, max_returns):
    company_name = company_being_read.offset(0, -1).value
    isin = company_being_read.value
    ia_sector = company_being_read.offset(0, 1).value
    average_cap = company_being_read.offset(0, 2).value
    g_v_s = company_being_read.offset(0, 3).value
    ucr = company_being_read.offset(0, 4).value
    dcr = company_being_read.offset(0, 5).value
    if dcr is None:
        dcr = 0
    if ucr is None:
        ucr = 0
    year_three = amr[0]
    year_five = amr[1]
    if amount_returns < max_returns * 0.5:
        return None
    # difference = 0
    # if dcr is not None and ucr is not None:
    #     difference = ucr - dcr

    return [company_name, isin, ia_sector, average_cap, g_v_s, year_three, year_five, amount_returns, ucr, dcr,
            ucr - dcr]


def generator(fund_dictionary):
    to_return = []
    i = 0
    for x in fund_dictionary.values():
        if len(x) > i:
            i = len(x)
    for key in fund_dictionary.keys():
        if fund_dictionary[key][-1].value is None:
            continue
        else:
            fund_dictionary[key] = [x for x in fund_dictionary[key] if x.value if not None]
        amr, three_y, five_y = process_weekly_returns(fund_dictionary[key])
        amr_returns = amr_calculator(amr, three_y, five_y)
        structured_output = output_generator(key, amr_returns, len(fund_dictionary[key]), i)
        if structured_output is not None:
            to_return.append(structured_output)
    else:
        return to_return


# This generates the deciles used for catagorisation.
def classify_deciles(output_ready_for_processing):
    third_year_rrr = []
    fifth_year_rrr = []
    for fund in output_ready_for_processing:
        if fund[4] is not None:
            third_year_rrr.append(fund[4])
        if fund[5] is not None:
            fifth_year_rrr.append(fund[5])
    third_year_deciles = np.quantile(third_year_rrr, q=np.arange(start=0.1, stop=1, step=0.1))
    fifth_year_deciles = np.quantile(fifth_year_rrr, q=np.arange(start=0.1, stop=1, step=0.1))
    return third_year_deciles, fifth_year_deciles


# This breaks the funds into their deciles using a function that is built into numpy.
def decimate_rrr(deciles, all_funds):
    for fund in all_funds:
        fund[5] = np.searchsorted(deciles[0], fund[5]) + 1
        fund[6] = np.searchsorted(deciles[1], fund[6]) + 1
    return all_funds


# There are 3 groups that the value growth scores are to be broken into.
# def find_value_growth_score_sheets(fund_array):
#     to_return = []
#     for fund in fund_array:
#         if fund[3] is None:
#             fund[3] = 0
#         to_return.append(fund[3])
#     return np.quantile(to_return, q=np.arange(0, 1, 1/3))[1:]


# Groups all funds by their growth/value scores.
def sort_growth_value_scores(funds, boundaries, headers):
    value, balanced, growth = [], [], []
    for i, head in enumerate(headers):
        if "value-growth" in head.lower():
            i = i - 1
    for fund in funds:
        print(fund)
        # TODO: Make this dynamic
        if fund[4] < boundaries[0]:
            value.append(fund)
        elif boundaries[0] < fund[4] < boundaries[1]:
            balanced.append(fund)
        else:
            growth.append(fund)
    return [value, balanced, growth]


# TODO: Microcap - Mid & Small breakdown. Take from Strip by RRR. micro's are less than between 200-300 see what works.
#
def high_cap_low_cap(grow_bal_val_sorted):
    value_low_cap, value_high_cap, balanced_low_cap, balanced_high_cap, growth_low_cap, growth_high_cap, = [], [], \
        [], [], \
        [], []
    for idx, v_g_s in enumerate(grow_bal_val_sorted):
        # TODO: What'd be good is having in here something that dynamically finds the location. Maybe I can assume
        #  that they'll be strings?
        for fund in v_g_s:
            ref1 = 0
            for i in range(0, 10):
                try:
                    ref1 = fund[i]
                    ref1 = float(ref1 + 1)
                    ref1 = i
                    break
                except TypeError:
                    pass
            if fund[ref1] is None:
                fund[ref1] = 0
            if fund[ref1] > 6000:
                if idx == 0:
                    value_high_cap.append(fund)
                elif idx == 1:
                    balanced_high_cap.append(fund)
                elif idx == 2:
                    growth_high_cap.append(fund)
            # elif fund[2] < 250:
            # TODO: This can be added to output at later time.
            # continue
            else:
                if idx == 0:
                    value_low_cap.append(fund)
                elif idx == 1:
                    balanced_low_cap.append(fund)
                elif idx == 2:
                    growth_low_cap.append(fund)

    return [value_low_cap, value_high_cap, balanced_low_cap, balanced_high_cap, growth_low_cap, growth_high_cap]


def get_headers():
    allowable_headers = []
    for i in range(1, 50):
        cell = READER_SHEET.cell(9, i).value

        # TODO: break in here if return create list of values
        if "return" not in cell.lower() or "cumulative" not in cell.lower():
            allowable_headers.append(cell)
        else:
            return allowable_headers


def main():
    global READER_SHEET, MODIFIED_FILENAME, ALL_FUNDS, ALL_RETURNS
    all_files = file_id_reader.list_all_files(True)
    filename_user = file_id_reader.user_selected_file(all_files)
    filenames = file_id_reader.file_name_generator(filename_user)
    for filename in filenames:
        reader_doc = xl.load_workbook(filename)
        READER_SHEET = reader_doc.active
        file_id_reader.READER_SHEET = READER_SHEET
        headers = get_headers()
        ALL_FUNDS = file_id_reader.find_all_funds()
        ALL_RETURNS = file_id_reader.returns_column()
        funds_dict = dictionary_creation()
        output_ready_for_processing = generator(funds_dict)
        print(output_ready_for_processing)
        rrr_deciles = classify_deciles(output_ready_for_processing)
        deciled_output = decimate_rrr(rrr_deciles, output_ready_for_processing)
        deciled_output = sorted(deciled_output, key=itemgetter(5), reverse=True)
        # This can be used as an approximate hard limit of value and growth - less than 100 is deepish value,
        # over 250 is highish growth
        grow_bal_val_boundaries = [100, 200]
        # This can replace the fixed boundaries above and takes the averages to give relatively equal quantity of
        # funds in each boundary
        # grow_bal_val_boundaries = find_value_growth_score_sheets(deciled_output)
        grow_bal_val_sorted = sort_growth_value_scores(deciled_output, grow_bal_val_boundaries, headers)
        print(grow_bal_val_sorted)
        fully_sorted_companies = high_cap_low_cap(grow_bal_val_sorted)
        file_id_reader.excel_writer(fully_sorted_companies, filename[:-5] + " RRR Calculation.xlsx")
    else:
        print("Successfully written")


if __name__ == '__main__':
    main()
