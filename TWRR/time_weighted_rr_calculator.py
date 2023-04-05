# TODO: Add function to get value of portfolio at user given date. Optionally, add function for last transaction to
#  be counted.
# TODO: This should really work with xlsxwriter whihc would mean I can insert the graph easier.
from datetime import datetime

import copy
import sys
import time
import logging

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

import pandas
import io

import scipy as sc


from operator import itemgetter
from numpy import ndarray

from Resources import file_id_reader

global TRANSACTIONS_SHEET, TRANSACTION_DATES, PRICES_SHEET, FUNDS_HELD, TRANSACTION_FILENAME
FIRST_DATE_INDEX = None
UNIT_TIME_SERIES = []

logging.basicConfig(filename="LoggingFile.log", level=logging.ERROR,
                    format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')


# This checks whether the transaction should be considered. This function should maybe be tweaked so that all transactions
# Are recorded, but not all transactions create holding periods. #
def deal_type_col_finder():
    for i in range(1, TRANSACTIONS_SHEET.max_column):
        cell = TRANSACTIONS_SHEET.cell(1, i)
        if cell.value == 'TradeType':
            return cell.column
    else:
        sys.exit("There does not appear to be a trade_type found.")


def check_transaction_type(transaction):
    deal_type = TRANSACTIONS_SHEET.cell(row=transaction.row, column=deal_type_col_finder())
    # I don't think we want this here, we need the transaction in to give a realistic report of holding return.
    if deal_type == "To be Ignored":
        return False
    else:
        return True


def get_all_transactions():
    """
    Find all transactions on given sheet.
    :return: Dates of all transactions
    """
    all_transactions = []
    for row_id in range(TRANSACTIONS_SHEET.max_row, 1, -1):
        cell = TRANSACTIONS_SHEET.cell(row=row_id, column=1)
        if check_transaction_type(cell):
            all_transactions.append(date_conversion(cell.value))
    else:
        return all_transactions


def date_conversion(date_str):
    """
    Takes a string in the format of a date and formats it.
    :param date_str: String
    :return: Date
    """
    try:
        date_str = datetime.strptime(date_str, '%d/%m/%Y')
        return date_str
    except TypeError:
        date_str = date_str.strftime('%d/%m/%Y')
        date_str = datetime.strptime(date_str, '%d/%m/%Y')
        return date_str


def remove_duplicates(transaction_list):
    """
    By converting a list to keys in a dictionary, it removes all duplicates in a list.
    :param transaction_list: Array
    :return: Array w/out duplicates.
    """
    return list(dict.fromkeys(transaction_list))


def get_transaction_dates():
    """
    Find all transaction dates in a portfolio.
    This can be used to determine the units at any point.
    As well as the total amount of transactions if all dates are to be used.
    :return: Array
    """
    global TRANSACTIONS_SHEET
    TRANSACTIONS_SHEET = get_transaction_sheet()
    transaction_list = get_all_transactions()
    no_date_duplicates = remove_duplicates(transaction_list)
    return no_date_duplicates


def get_transaction_sheet():
    global TRANSACTION_FILENAME
    all_file_list = file_id_reader.list_all_files(True, '.xlsx')
    transactions_filename = file_id_reader.user_selected_file(all_file_list)
    TRANSACTION_FILENAME = transactions_filename
    transactions_file = openpyxl.load_workbook(transactions_filename, data_only=True)
    return transactions_file.active


def find_fund_column():
    for i in range(1, TRANSACTIONS_SHEET.max_column):
        cell = TRANSACTIONS_SHEET.cell(row=1, column=i)
        if cell.value == "Share_Class_Name_Long":
            return cell.column


def list_of_funds_generator(fund_column):
    fund_list = []
    for i in range(2, TRANSACTIONS_SHEET.max_row + 1):
        cell = TRANSACTIONS_SHEET.cell(row=i, column=fund_column)
        fund_list.append(cell.value)
    else:
        return fund_list


def dictionary_creator(tidy_fund_list):
    return_dictionary = {}
    for i in tidy_fund_list:
        return_dictionary[i] = 0
    return return_dictionary


def units_column_finder():
    for col in range(1, TRANSACTIONS_SHEET.max_column):
        cell = TRANSACTIONS_SHEET.cell(row=1, column=col)
        if cell.value is not None and cell.value.lower() == 'units':
            return cell.column


def consideration_col_finder():
    for col in range(1, TRANSACTIONS_SHEET.max_column):
        cell = TRANSACTIONS_SHEET.cell(row=1, column=col)
        if cell.value is not None and cell.value.lower() == 'consideration':
            return cell.column


def cost_to_holder_col_finder():
    for col in range(1, TRANSACTIONS_SHEET.max_column):
        cell = TRANSACTIONS_SHEET.cell(row=1, column=col)
        if cell.value is not None and cell.value.lower() == 'costtoholder':
            return cell.column


def unit_finder(fund_dictionary, transaction_date):
    fund_dictionary['Cost_to_holder'] = 0
    fund_dictionary['Transaction_Costs'] = 0
    fund_dictionary['TradeType'] = str()
    units_col = units_column_finder()
    fund_col = find_fund_column()
    trade_type_col = deal_type_col_finder()
    consideration_col = consideration_col_finder()
    c_t_h_col = cost_to_holder_col_finder()
    for row in range(2, TRANSACTIONS_SHEET.max_row + 1):
        cell = TRANSACTIONS_SHEET.cell(row=row, column=1)
        cell_val = cell.value
        cell_date = date_conversion(cell_val)
        if cell_date == transaction_date:
            fund_name = TRANSACTIONS_SHEET.cell(row=row, column=fund_col).value
            units = TRANSACTIONS_SHEET.cell(row=row, column=units_col)
            if fund_name in fund_dictionary:
                fund_dictionary[fund_name] += units.value
                if fund_dictionary[fund_name] < 1:
                    fund_dictionary[fund_name] = 0
                fund_dictionary['TradeType'] = TRANSACTIONS_SHEET.cell(row, trade_type_col).value
                try:
                    fund_dictionary['Transaction_Costs'] += TRANSACTIONS_SHEET.cell(row, c_t_h_col).value
                except KeyError:
                    fund_dictionary['Transaction_Costs'] = TRANSACTIONS_SHEET.cell(row, c_t_h_col).value

                try:
                    fund_dictionary['Cost_to_holder'] += TRANSACTIONS_SHEET.cell(row, consideration_col).value
                except KeyError:
                    fund_dictionary['Cost_to_holder'] = TRANSACTIONS_SHEET.cell(row, consideration_col).value
    else:
        return fund_dictionary


def all_funds_held():
    fund_column = find_fund_column()
    find_aggregate_list_of_funds = list_of_funds_generator(fund_column)
    tidy_fund_list = remove_duplicates(find_aggregate_list_of_funds)
    fund_dictionary = dictionary_creator(tidy_fund_list)
    return unit_finder(fund_dictionary, TRANSACTION_DATES[0])


def get_prices_sheet():
    global PRICES_SHEET
    prices_filename = file_id_reader.user_selected_file(file_id_reader.list_all_files(True, '.csv'), "prices")
    pandas_pricing = pd.read_csv(prices_filename, dtype=str, usecols=[0, 1, 2, 3, 4])
    pandas_pricing.dropna(inplace=True)
    PRICES_SHEET = np.array(pandas_pricing)


# Here we will attempt to create a list of dictionaries.
def unit_series_calculator():
    global FUNDS_HELD
    new_array = []
    blank_dict = {}
    for date in range(1, len(TRANSACTION_DATES)):
        blank_dict = FUNDS_HELD.copy()
        new_array.append(blank_dict)
        FUNDS_HELD = unit_finder(FUNDS_HELD, TRANSACTION_DATES[date])
        blank_dict = FUNDS_HELD.copy()
    else:
        new_array.append(blank_dict)
    return new_array


def sum_all_hps(time_series):
    dict_to_return = {}
    transaction_costs, cost_to_holder, trade_type = [], [], []
    for idx, holding_period in enumerate(time_series):
        try:
            transaction_costs.append(holding_period.pop("Transaction_Costs"))
        except KeyError:
            holding_period["Transaction_Costs"] = 0
            transaction_costs.append(holding_period.pop("Transaction_Costs"))
        try:
            cost_to_holder.append(holding_period.pop("Cost_to_holder"))
        except KeyError:
            holding_period["Cost_to_holder"] = 0
            cost_to_holder.append(holding_period.pop("Cost_to_holder"))
        trade_type.append(holding_period.pop("TradeType"))
        dict_to_return[TRANSACTION_DATES[idx]] = sum(holding_period.values())
    else:
        return dict_to_return, transaction_costs, cost_to_holder, trade_type


def convert_units_to_values(units):
    to_ignore = ['cost_to_holder', 'transaction_costs', 'trade_type']
    for idx, holding_period in enumerate(units):
        transaction_date = TRANSACTION_DATES[idx]
        res = (binary_search(PRICES_SHEET, 0, len(PRICES_SHEET) - 1, TRANSACTION_DATES[idx]))
        if res[1]:
            for key, val in holding_period.items():
                if not key.lower() in to_ignore:
                    # Because there are multiple dates contained within the prices sheet and the binary search only finds
                    # the index of one, somewhere in this list, this range is given to allow for the selected cell to be
                    # 200 away. Whilst creating a function that checks for each date span, and dynamically presents a range
                    # would be elegant. Computationally it is adding another step in processing, rather than this assumption
                    # That can run through the data (because it's in numpy) very quickly.
                    # If this is not suitably elegant, adding a variable called 'total_share_classes' that would allow
                    # easier modification could be another option. #
                    # print(transaction_date)
                    for fund_prices in range(res[1] - 200, res[1] + 200):
                        try:
                            valuation_date = datetime.strptime(PRICES_SHEET[fund_prices][3], '%d/%m/%Y')
                            if valuation_date == transaction_date:
                                if PRICES_SHEET[fund_prices][2].lower() == key.lower():
                                    holding_period[key] = val * (float(PRICES_SHEET[fund_prices][4]) / 100)
                        # These are here in case there is unexpected behaviour. This is all errors that're caught
                        #   The benefit of this is that they can be reintroduced very simply.
                        except AttributeError as e:
                            print(e)
                        except TypeError as e:
                            print(e)
                        except ValueError as e:
                            print(e)
                        except IndexError:
                            pass
    return units


# This is a custom binary search function. Praise me.
def binary_search(array, low, high, x):
    if high >= low:
        mid_point = (high + low) // 2
        try:
            arr_date = datetime.strptime(array[mid_point][3], "%d/%m/%Y")
        except ValueError:
            arr_date = datetime.strptime(array[mid_point+1][3], "%d/%m/%Y")
        if x == arr_date:
            return array[mid_point], mid_point
        elif arr_date > x:
            return binary_search(array, low, mid_point - 1, x)
        elif arr_date < x:
            return binary_search(array, mid_point + 1, high, x)
    return -1, False


def user_date_getter():
    global TRANSACTION_DATES
    while True:
        try:
            print("Please enter date in format %d/%m/%Y\nEnter 'all' for all transactions listed.")
            user_start_str = input("Enter the start date: ")
            if user_start_str.lower() == "all":
                return 'all', True
            user_end_str = input("Enter the end date: ")
            user_start_date = date_conversion(user_start_str)
            user_end_date = date_conversion(user_end_str)
            if user_start_date < user_end_date:
                return user_start_date, user_end_date
            else:
                print("The dates given are not in the right order")
        except Exception as e:
            print(e)


def user_transactions_getter(user_start_date, user_end_date, unit_time_series):
    global TRANSACTION_DATES
    # If transaction dates are to be added, rather than all transactions to be used, then this will require further
    #   Development
    if user_start_date.lower() == "all":
        return unit_time_series

        # except AttributeError as e:
        #     logging.error(e, exc_info=True)


        # transaction_date = date_conversion(date)
        # if idx == len(TRANSACTION_DATES) - 1:
        #     print("THIS IS HIT 1")
        #     user_requested_dates.append(date)
        #     user_requested_values.append(unit_time_series[idx])
        #     break
        #
        # if user_start_date < transaction_date and idx == 0:
        #     sys.exit(f"This means that the date requested as the start date: {user_start_date} was before the first "
        #              f"transaction recorded, which is: {date}")
        # if user_start_date <= transaction_date <= user_end_date:
        #     print("THIS IS HIT 2")
        #     user_requested_dates.append(date)
        #     user_requested_values.append(unit_time_series[idx])
        # try:
        #     if transaction_date < user_end_date < TRANSACTION_DATES[idx + 1]:
        #         print("THIS IS HIT 3")
        #         user_requested_dates.append(user_end_date)
        #         user_requested_values.append(unit_time_series[idx - 1])
        #         break
        # except IndexError:
        #     logging.error(IndexError, exc_info=True)
        # if transaction_date < user_start_date:
        #     try:
        #         print("THIS IS HIT 4")
        #         user_requested_dates[0] = user_start_date
        #         user_requested_values[0] = unit_time_series[idx]
        #     except IndexError as ie:
        #         logging.error(ie, exc_info=True)
        #         user_requested_dates.append(user_start_date)
        #         user_requested_values.append(unit_time_series[idx])
        # if user_end_date < transaction_date:
        #     print("THIS IS HIT 6")
        #     user_requested_dates.append(date)
        #     user_requested_values.append(unit_time_series[idx])
        #     break
    # user_requested_values, user_requested_dates = [], []
    # for idx, date in enumerate(TRANSACTION_DATES):
    #     try:
    # TRANSACTION_DATES = user_requested_dates
    # return user_requested_values


def main():
    global TRANSACTION_DATES, FUNDS_HELD
    TRANSACTION_DATES = get_transaction_dates()
    get_prices_sheet()
    FUNDS_HELD = all_funds_held()

    user_start_date, user_end_date = user_date_getter()

    unit_time_series = unit_series_calculator()
    user_requested_dates = user_transactions_getter(user_start_date, user_end_date, unit_time_series)

    user_units_to_process = copy.deepcopy(user_requested_dates)
    value_time_series = convert_units_to_values(user_units_to_process)

    summed_holding_periods, transaction_costs, cost_to_holder, trade_type = sum_all_hps(value_time_series)

    portfolio_valuations = []
    cost_to_holder_arr = []
    date_arr = []
    for idx, (holding_date, holding_value) in enumerate(summed_holding_periods.items()):
        portfolio_valuations.append(holding_value)
        date_arr.append(TRANSACTION_DATES[idx])
        cost_to_holder_arr.append(cost_to_holder[idx])

    return_percents = []
    for idx, valuation in enumerate(portfolio_valuations):
        if idx != len(portfolio_valuations) - 1:
            return_val = (portfolio_valuations[idx + 1] - (
                    portfolio_valuations[idx] + (cost_to_holder_arr[idx + 1]))) / (
                                 portfolio_valuations[idx] + (cost_to_holder_arr[idx + 1]))
            return_percents.append(return_val)
    list_twrr, twrr = [0], 1
    for idx, hp_ret in enumerate(return_percents):
        twrr = ((1 + hp_ret) * twrr)
        list_twrr.append((twrr - 1) * 100)
    else:
        master_list, units_to_output = [], []

    print(f"Identified {len(list_twrr)} holding periods.")

    aggregate_total_invested = 0
    aggregate_cost_to_holder = 0
    # Array creation to parse into DataFrames
    for idx in range(len(list_twrr)):
        tran_date = date_arr[idx]
        tran_date = tran_date.strftime('%d/%m/%Y')
        port_val = str(round(portfolio_valuations[idx], 2))
        twrr_to_write = str(round(list_twrr[idx], 2)) + "%"
        units_to_write = []
        c_t_h = str(round(unit_time_series[idx]['Cost_to_holder'], 2))
        trade_type = unit_time_series[idx]['TradeType']
        if idx == 0:
            return_of_hp = str('0%')
        else:
            return_of_hp = '{0}%'.format(str(f'{round(return_percents[idx - 1] * float(100), 2)}'))
        quantity_of_funds = 0
        for key, val in unit_time_series[idx].items():
            if key != 'Cost_to_holder' and key != 'Transaction_Costs' and key != 'TradeType':
                units_to_write.append(val)
                quantity_of_funds += 1
        if unit_time_series[idx]['TradeType'] == 'Periodic Adviser Charge':
            aggregate_cost_to_holder += abs(unit_time_series[idx]['Cost_to_holder'])
        if "investment" in unit_time_series[idx]['TradeType'].lower():
            aggregate_total_invested += abs(unit_time_series[idx]['Transaction_Costs'])
        master_list.append([tran_date, trade_type, c_t_h, port_val, twrr_to_write, str(return_of_hp)])
        units_to_output.append(units_to_write)
    else:
        fund_names = [f"{i} Units" for i in unit_time_series[1].keys()]

    print(f"Total invested: {aggregate_total_invested}. Total costs: {aggregate_cost_to_holder}")

    # Dataframe generation for writing.
    output_filename = TRANSACTION_FILENAME[:-5] + " Output.xlsx"
    # Transaction History Sheet
    df1 = pd.DataFrame(master_list,
                       columns=['Transaction Date', 'Transaction Type', 'Cost to Holder', 'Portfolio Valuation',
                                'Cumulative TWR', 'TWR HP'])
    df2 = pd.DataFrame(units_to_output, columns=fund_names[:-3])
    df3_data = [str(round(list_twrr[len(list_twrr) - 1], 2)) + '%', str(aggregate_total_invested),
                str(round(portfolio_valuations[len(portfolio_valuations) - 1], 2)),
                str(aggregate_cost_to_holder)]
    # Summary sheet
    df3_cols = ["Total Portfolio Return", "Total Amount Invested", "End Portfolio Value",
                "Charges Deducted"]
    df3 = pd.DataFrame(df3_data).T
    df3.columns = df3_cols
    df4 = pd.concat([df1, df2], axis=1)
    with pd.ExcelWriter(output_filename) as writer:
        df4.to_excel(writer, index=False, sheet_name="Transaction History")
        df3.to_excel(writer, index=False, sheet_name="Summary Tab")
        print("Dataframe written to Excel successfully.")

    # Convert arrays to numpy arrays for plotting
    twrr_numpy = np.asarray(list_twrr)
    port_values_numpy = np.asarray(portfolio_valuations)
    date_numpy = np.asarray(date_arr)

    # Generete first plot, and left-hand y-axis.
    fig, ax = plt.subplots()
    ax.set_xlabel("Date")
    ax.set_ylabel("TWRR As %", color="red")
    ax.plot(date_numpy, twrr_numpy, color="red")
    ax.tick_params(axis='y', labelcolor='red')

    ax2 = ax.twinx()
    # Onto the same plot, generate right-hand y-axis.
    ax2.set_ylabel('Cumulative Portfolio Value', color='blue')
    ax2.plot(date_numpy, port_values_numpy, color="blue")
    ax2.tick_params(axis='y', labelcolor='blue')

    # Ensure that labels display on plot.
    plt.tight_layout()
    fig.autofmt_xdate()

    # Label x-axis and display the graph.
    plt.savefig("TWRR & Portfolio Graph.png")
    plt.close()
    print("Successfully generated the graph.")


if __name__ == '__main__':
    main()
    time.sleep(0.1)
    sys.exit("Output file and graph successfully generated. Program will now close.")
