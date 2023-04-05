import os
import sys
import openpyxl as xl
import pandas as pd

# Variable initalisation.
global READER_SHEET, MODIFIED_FILENAME

RRR_HEADERS = ["Fund Name", "ISIN", "Average Cap", "Growth to Value Score", "3 Year RRR", "5 Year RRR",
               "Amount of periods utilised", "Down Capture Ratio", "Up Capture Ratio", "Up Down Difference"]


# This lists all xlsx files which can be read by the rebalancer.
def list_all_files(to_print=True, file_type='.xlsx', exclusions: list = None):
    """
    List all files according to certain criteria.
    :param exclusions: This is a list that takes files that should be excluded. This is going to be for test files,
    templates, that kind of thing.
    :params:
    :param to_print: Boolean
    | Is this function to print out all files found. Can be used when the same list would be printed twice.
    | This helps avoid clutter in the console.
    :param file_type: String
    | The filetype that is to be listed. This can be passed an empty string to return all files in directory.
    :return: Array
    """
    a = 1
    files_ = []
    for x in os.listdir():
        if (x.endswith(file_type) or x.endswith(file_type)) and not x.startswith('~'):
            if x not in exclusions:
                if to_print:
                    print(a, x)
                files_.append(x)
                a += 1
    return files_


def user_selected_file(file_list, doc_type='file', enable_all=True, default_return=False):
    """
    Allows the user to return an index for the desired file. doctype allows for printing various file_types.
    Returns the filename to.
    :param default_return:
    :param enable_all:
    :param file_list: Array
    :param doc_type: String
    :return: String
    """
    while True:
        if default_return:
            return file_list[0]
        desired_index = input(f"Input the index of the {doc_type} you'd like to read {'enter a for all files' if enable_all else ''}:")
        try:
            desired_index = int(desired_index)
            if len(file_list) + 1 > desired_index > 0:
                break
        except ValueError:
            if desired_index == 'a':
                return 'a'
            elif desired_index.lower() == 'exit':
                sys.exit("Quitting")
            print("NaN")
    return file_list[int(desired_index) - 1]


def file_name_generator(filename):
    """
    Generates all filenames required
    :param filename: String
    :return: Funct / String
    """
    if filename == 'a':
        return list_all_files(False)
    else:
        return [filename]


def sheet_reader(filename, data_only=True, sheet_name='Rebalancer'):
    document = xl.load_workbook(filename, data_only=data_only)
    if sheet_name == "Active":
        return document.active
    else:
        return document[sheet_name]


def find_all_funds():
    funds_started = False
    all_funds = []
    for i in range(1, READER_SHEET.max_row):
        isin = READER_SHEET.cell(i, 2)
        av_cap = READER_SHEET.cell(i, 3).value
        cal_gro = READER_SHEET.cell(i, 4).value
        if funds_started and isin.value is not None and av_cap is not None and cal_gro is not None:
            all_funds.append(isin)
        if isin.value is not None and isin.value.lower() == 'isin':
            funds_started = True
    else:
        return all_funds


def returns_column():
    returns_started = False
    return_weeks = []
    for i in range(1, READER_SHEET.max_column):
        cell = READER_SHEET.cell(9, i)
        if 'return' in cell.value.lower() and returns_started is False:
            returns_started = True
        elif 'information' in cell.value.lower():
            return_weeks = return_weeks[:-1]
            return return_weeks
        if returns_started:
            return_weeks.append(cell)
    else:
        return return_weeks


def five_year_finder_regression():
    for i in range(1, READER_SHEET.max_column):
        if "ROE" in READER_SHEET.cell(9, i).value:
            return i + 1


def excel_writer(fully_sorted_companies, modified_filename):
    with pd.ExcelWriter(modified_filename, engine="openpyxl") as writer:
        # So, create an array in here of the value small value large etc, then just iterate through fully sorted
        # companies. This can then have a if len(fullysorted[i]) != 0 write.
        pd.DataFrame(fully_sorted_companies[0], columns=RRR_HEADERS).to_excel(writer, sheet_name="Value Small Cap",
                                                                              index=False)
        pd.DataFrame(fully_sorted_companies[1], columns=RRR_HEADERS).to_excel(writer, sheet_name="Value Large Cap",
                                                                              index=False)
        pd.DataFrame(fully_sorted_companies[2], columns=RRR_HEADERS).to_excel(writer, sheet_name="Balanced Small Cap",
                                                                              index=False)
        pd.DataFrame(fully_sorted_companies[3], columns=RRR_HEADERS).to_excel(writer, sheet_name="Balanced Large Cap",
                                                                              index=False)
        pd.DataFrame(fully_sorted_companies[4], columns=RRR_HEADERS).to_excel(writer, sheet_name="Growth Small Cap",
                                                                              index=False)
        pd.DataFrame(fully_sorted_companies[5], columns=RRR_HEADERS).to_excel(writer, sheet_name="Growth Large Cap",
                                                                              index=False)
