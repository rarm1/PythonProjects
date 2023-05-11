import datetime
import sys

import openpyxl as xl
import pandas as pd

from assets import file_id_reader
import variables
# from Introduction import Intro

# TODO: Does the 5, 3, 1 and YTD comparisons need to have seperate thresholds for outperformance.
#  I feel like they do because they are different time periods.
# TODO: Portfolio concentration does not populate.
# TODO: Over a year to date gives the wrong month. See why this is the case. Do I just use a fixed point in the
#  template for this? Could that be automated?
# TODO: All of the percentages that're displayed with .0% on the end shouldn't be displayed.
# TODO: Bring this into classes separate it out to make it more modular and easy to update.
# TODO: When doing so bring the templates into Python. It is clumsy and hard to maintain when it is built into Excel
#  templates.
# TODO: Strategies perform 'approximately' in line. Maybe add thresholds for this too.
# todo: Can we superscript things like dates?
global DOCUMENT, VAR_DICT, FILENAME


def obtain_file():
	"""

	:return:
	"""
	global FILENAME
	all_files = file_id_reader.list_all_files(exclusions=["Comparison Template.xlsx"])
	filename_user = file_id_reader.user_selected_file(all_files)
	filename = file_id_reader.file_name_generator(filename_user)[0]
	FILENAME = filename
	port_hold_df = pd.read_excel(filename, sheet_name="Portfolio holdings", header=None)
	port_comp_df = pd.read_excel(filename, sheet_name="Portfolio Comparison", header=None)
	perf_comp_df = pd.read_excel(filename, sheet_name="Performance Comparison", header=None)
	document = xl.load_workbook(filename, data_only=True)
	return port_hold_df, port_comp_df, perf_comp_df


def variable_definitions():
	"""

	"""
	global VAR_DICT
	var_sheet = DOCUMENT['Variables']
	variable_cell_list = [var_sheet.cell(row=a, column=1) for a in range(2, var_sheet.max_row + 1) if var_sheet.cell(
			row=a, column=1).value is not None]
	VAR_DICT = {a.value: a.offset(0, 1).value for a in variable_cell_list}


def harpoon():
	"""

	"""
	global DOCUMENT
	port_hold_df, port_comp_df, perf_comp_df = obtain_file()
	intro = variables.Intro(port_hold_df, port_comp_df, perf_comp_df)
	print(intro.intro_text)
	asset_allocation = variables.AssetAllocation(port_hold_df, port_comp_df, perf_comp_df)
	print(asset_allocation.asset_allocation_text)

	# variable_definitions()
	# written_row = 1
	# output_sheet = DOCUMENT.create_sheet(title="OUTPUT")
	# # TODO: Create classes out of these sheets.
	#
	# input_sheets = ["Introduction", "Asset Allocation", "Performance", "Ongoing Charges"]
	# for input_sheet in input_sheets:
	# 	paragraphs = paragraph_writer(input_sheet)
	# 	output_sheet.cell(row=written_row, column=1).value = input_sheet
	# 	written_row += 1
	# 	for para in paragraphs:
	# 		output_sheet.cell(row=written_row, column=1).value = para
	# 		written_row += 1
	# port_comp_sheet = DOCUMENT['Portfolio Comparison']
	# margetts_ocf, client_ocf, client_name, margetts_fund = 0.0, 0.0, str, []
	# for i in range(1, 30):
	# 	try:
	# 		if 'Margetts Weighted OCF' in port_comp_sheet.cell(i, 7).value:
	# 			margetts_ocf = port_comp_sheet.cell(i, 8).value
	# 		if 'Weighted Average Client OCF' in port_comp_sheet.cell(i, 7).value:
	# 			client_ocf = port_comp_sheet.cell(i, 8).value
	# 		if 'Margetts Fund' in port_comp_sheet.cell(i, 7).value:
	# 			margetts_fund.append(port_comp_sheet.cell(i, 8).value)
	# 		if 'Margetts Share Class' in port_comp_sheet.cell(i, 7).value:
	# 			margetts_fund.append(port_comp_sheet.cell(i, 8).value)
	# 		if 'Client Name' in port_comp_sheet.cell(i, 7).value:
	# 			client_name = port_comp_sheet.cell(i, 8).value
	# 	except TypeError:
	# 		# Because of the 'in' clause, this will error. Could be replaced with == when template is fully resolved.
	# 		pass
	# margetts_fund = " ".join(margetts_fund)
	# output_sheet.cell(row=written_row, column=1).value = "Margetts Risk Rated " + margetts_fund
	# output_sheet.cell(row=written_row, column=2).value = str(round(margetts_ocf * 100, 2)) + "%"
	# written_row += 1
	# output_sheet.cell(row=written_row, column=1).value = client_name + " portfolio (weighted average underlying OCF)"
	# output_sheet.cell(row=written_row, column=2).value = str(round(client_ocf * 100, 2)) + "%"
	# DOCUMENT.save(FILENAME[:-5] + " Output.xlsx")


def higher_lower_calc(var):
	"""

	:param var:
	:return:
	"""
	margetts_weight, clients_weight = VAR_DICT["margetts" + var[var.find('_'):-13]], VAR_DICT["client" + var[var.find(
			'_'):-13]]
	var1, var2 = [clients_weight, margetts_weight] if 'client' in var.lower() else [margetts_weight, clients_weight]
	diff = var1 - var2
	if abs(diff) > 0.02:
		return "higher" if diff > 0 else "lower"
	else:
		return "similar"


def annualised_calc():
	"""

	:return:
	"""
	perf_comp = DOCUMENT['Performance Comparison']
	target_cells = []
	start = False
	for a in range(1, perf_comp.max_row + 1):
		if start:
			year = perf_comp.cell(a, 1).value
			margetts = round(perf_comp.cell(a, 2).value, 4)
			client = round(perf_comp.cell(a, 3).value, 4)
			target_cells.append([year, margetts, client])
		if perf_comp.cell(a, 1).value == "Annualised":
			start = True
	
	margetts_outperf_years = [str(a) for a, b, c in target_cells if b - c > 0.01]
	if len(margetts_outperf_years) >= 2:
		margetts_outperf_years[-1] = "and " + margetts_outperf_years[-1]
	margetts_outperf_years_str = ", ".join(margetts_outperf_years)
	client_outperf_years = [str(a) for a, b, c in target_cells if b - c < -0.01]
	if len(client_outperf_years) >= 2:
		client_outperf_years[-1] = "and " + client_outperf_years[-1]
	client_outperf_years_str = ", ".join(client_outperf_years)
	equal_years = [str(a) for a, b, c in target_cells if 0.01 >= b - c >= -0.01]
	if len(equal_years) >= 2:
		equal_years[-1] = "and " + equal_years[-1]
	equal_years_str = ", ".join(equal_years)
	to_return = ["The below chart shows discrete annual performance for the last 5 years."]
	if len(margetts_outperf_years) > 0:
		rr_outperformance = f"The Margetts Risk Rated {VAR_DICT['margetts_portfolio_identifier']} model, " \
		                    f"performed ahead of {VAR_DICT['client_name']}'s portfolio across " \
		                    f"{margetts_outperf_years_str}."
		to_return.append(rr_outperformance)
	if len(client_outperf_years) > 0:
		client_outperformance = f"{VAR_DICT['client_name']}'s portfolio outperformed the Margetts strategy in " \
		                        f"{client_outperf_years_str}."
		to_return.append(client_outperformance)
	if len(equal_years) > 0:
		in_line = f"The two portfolios performed in line across {equal_years_str}."
		to_return.append(in_line)
	to_return = " ".join(to_return)
	return to_return


# FIXME: if float(num) == float(int(num)): return int(num) else: return float(num)

def performance_reporting(variable_input):
	"""

	:param variable_input:
	:return:
	"""
	if "annualised_perf" in variable_input:
		return annualised_calc()
	elif "relative" not in variable_input:
		try:
			num = round(VAR_DICT[variable_input] * 100, 1)
			if float(num) == float(int(num)):
				num = int(num)
			num = str(num)
			return num
		except TypeError:
			print(f"variable_input: {variable_input}. Val: {VAR_DICT[variable_input]}")
	else:
		margetts_perf = VAR_DICT[variable_input[:4] + 'performance_margetts'] * 100
		client_perf = VAR_DICT[variable_input[:4] + 'performance_client'] * 100
		margetts_outperformance = margetts_perf - client_perf
		if margetts_outperformance >= 2:
			return 'ahead of '
		elif margetts_outperformance <= -2:
			return 'behind '
		elif 2 > margetts_outperformance > -2:
			# TODO: Maybe something like if abs outperformance is < 0.5 then it's in line else (so between 0.5 and 2)
			#  then it's approxiamtely in line.
			return 'approximately in line with '


def date_suffix(var):
	"""

	:param var:
	:return:
	"""
	dictionary = {
		'st': [1, 21, 31],
		'nd': [2, 22],
		'rd': [3, 23]
	}
	for key in dictionary.keys():
		if var in dictionary[key]:
			return key
	else:
		return "th"


def client_risk_comparison():
	"""

	:return:
	"""
	sheet = DOCUMENT['Portfolio Comparison']
	try:
		col_index = [i for i, cell in enumerate(sheet[1]) if cell.value == "Risk Score"][0]
		margetts_risk = sheet.cell(2, col_index + 1).value
		client_risk = sheet.cell(3, col_index + 1).value
	except Exception as e:
		sys.exit("The template is not working as expected. Please ensure Risk Score in on Portfolio Comparison sheet "
		         f"in top row. {e}")
	if margetts_risk - client_risk > 1:
		return "takes less "
	elif margetts_risk - client_risk < -1:
		return "takes more "
	else:
		return "takes a similar amount of "


def paragraph_writer(sheetname):
	"""

	:param sheetname:
	:return:
	"""
	intro_sheet = DOCUMENT[sheetname]
	paragraph = []
	for a, var in enumerate(intro_sheet.iter_rows(min_row=2, max_col=1, values_only=True), intro_sheet.max_row + 1):
		# var = intro_sheet.cell(row=a, column=1).value
		var = var[0]
		if var in VAR_DICT.keys():
			# Predetermined (client name, portfolio name, etc)
			if isinstance(VAR_DICT[var], str):
				paragraph.append(VAR_DICT[var])
			# Perf date, date of request
			elif isinstance(VAR_DICT[var], datetime.datetime):
				day = VAR_DICT[var].strftime("%d")
				paragraph.append(str(VAR_DICT[var].strftime(f"%d{date_suffix(int(day))} %B %Y")))
			# Performance figure written straight to document no processing
			elif isinstance(VAR_DICT[var], float) and "performance" not in var:
				perf_num = round(VAR_DICT[var] * 100, 1)
				if float(perf_num) == float(int(perf_num)):
					perf_num = int(perf_num)
				paragraph.append(str(perf_num))
			# Holding distribution comparison
			elif VAR_DICT[var] is None and "higher_lower" in var:
				paragraph.append(higher_lower_calc(var))
			elif "performance" in var:
				to_append = performance_reporting(var)
				paragraph.append(to_append)
			elif "client_portfolio_risk" in var:
				paragraph.append(client_risk_comparison())
		else:
			paragraph.append(var)
	paragraph = "".join(paragraph)
	paragraph = paragraph.split('new_line')
	return paragraph


if __name__ == '__main__':
	harpoon()
