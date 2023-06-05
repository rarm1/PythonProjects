import sys
from datetime import date
from os.path import exists

import openpyxl as xl
import pandas as pd

# Variable initialisation
global MODIFIED_FILENAME, WRITER_SHEET, READER_SHEET, READER_DOC, NON_BROKER_FILENAME, BROKER_FILENAME, FUND_MANAGER
TRADES = 0

SCHEME_LOOKUP_DOC = xl.load_workbook("X:\Fund Management\Fund Management Team Files\FM Personal Folders\Richard\Pycharm"
									 "Projects\Trade Reader\Resources\scheme_list_desig.xlsm")
SCHEME_SHEET = SCHEME_LOOKUP_DOC['Source']
ISIN_TO_TICKER_PATH = 'X:\Fund Management\Fund Management Team Files\FM Personal Folders\Richard\PycharmProjects\Trade Reader\Resources\ISIN_to_Ticker.csv'
ISIN_TO_TICKER_DOC = pd.read_csv(ISIN_TO_TICKER_PATH)
# ISIN_TO_TICKER_SHEET = ISIN_TO_TICKER_DOC['Sheet1']
FOLDER_DATE = date.today().strftime("%Y%m%d")
FUND_MANAGERS = {
	"Dmitry Konev": "DMITRY",
	"Richard Cole": "FMONEY",
	"Wayne Buttery": "OPES",
	"IBOSS Asset Management": "IBOSS",
	"Sentinel Asset Management": "SENTINEL"
}



def trade_array_generator(idx, trade):
	today = date.today()
	req_date = today.strftime("%d/%m/%Y")
	sec_name = trade.offset(0, -7).value
	isin = trade.offset(0, -9).value
	deal_type = _trade_type(trade)
	trade_value = _value_of_trade(deal_type, trade)
	trade_id = _get_trade_identifier() + idx
	notes = _notes_value(trade)
	return [FUND_MANAGER, trade_id, req_date, scheme_lookup().value, sec_name,	isin, new_buy_finder(trade),
			buy_or_sell(trade), deal_type, trade_value, currency_finder(trade), "N", notes, READER_SHEET.cell(2, 2).value,
			scheme_lookup().offset(0, 1).value]


def fundmanager_selection():
	global FUND_MANAGER
	for idx, x in enumerate(FUND_MANAGERS, start=1):
		print(f"{idx}: {x}")
	while True:
		desired_index = input("Please input the number of the fund manager to select: ")
		try:
			desired_index = int(desired_index)
			if len(FUND_MANAGERS) + 1 > desired_index > 0:
				break
		except ValueError:
			print("NaN")
	print("\n")
	FUND_MANAGER = list(FUND_MANAGERS.values())[desired_index - 1]


def scheme_lookup():
	rebalancer_scheme = READER_SHEET.cell(4, 2)
	for x in range(1, SCHEME_SHEET.max_row + 1):
		lookup_designation = (SCHEME_SHEET.cell(x, 3))
		if str(lookup_designation.value) in str(rebalancer_scheme.value) or str(rebalancer_scheme.value) in str(
				lookup_designation.value):
			return SCHEME_SHEET.cell(row=x, column=1)
	return f"Scheme PID {rebalancer_scheme.value} not listed in the scheme lookup doc."


def new_buy_finder(trade):
	if trade.offset(0, -2).value == 0:
		return "Y"
	else:
		return "N"


def buy_or_sell(trade):
	if trade.offset(0, 2).value > 0:
		return "BUY"
	else:
		return "SELL"


def _get_trade_identifier():
	trade_counter_doc = xl.load_workbook(
		"X:/Fund Management/Fund Management Team Files/FM Personal Folders/Richard/PycharmProjects/Trade "
		"Reader/Resources/trade_identifier.xlsx")
	trade_sheet = trade_counter_doc["trade_identifier"]
	return trade_sheet.cell(1, 1).value


def _value_of_trade(deal_type, trade):
	if "sellall" in deal_type.lower():
		return 0
	elif "cash" in deal_type.lower():
		return abs(trade.offset(0, 2).value)
	elif "unit" in deal_type.lower():
		return abs(trade.offset(0, 2).value)


def _trade_type(trade):
	holding_value = trade.offset(0, -6).value
	trade_value = trade.offset(0, 2).value
	if trade_value < 0:
		if _notes_value(trade) == "Sell All":
			if holding_value == abs(trade_value):
				return 'SellAll'
			else:
				print(READER_DOC)
				print(f"There has been an error with trade at coordinates: {trade}")
				input("Press any key to continue, the program will exit after this and allow for secondary checks.")
				sys.exit("THERE HAS BEEN A PROBLEM WITH SELL ALL'S NOT BEING NOTED IN NOTES AND/OR VALUATION"
				         "PLEASE RESOLVE THIS ISSUE BEFORE CONTINUING.")
	elif "broker" in _notes_value(trade).lower() or "etf" in _notes_value(trade).lower():
		return "Unit"
	return "Cash"


def sell_all_reader(trade):
	"""
	Determines whether it is a sell all or a normal rebalancing trade.
	:param trade: Cell containing trade
	:return: Returns either sell all or rebalancing
	"""
	if abs(round(trade.offset(0, 3).value, 2)) == 0.0:
		if _notes_value(trade) == "Sell All":
			return 'Sell All'
		else:
			print(READER_DOC)
			print(f"There has been an error with trade at coordinates: {trade}"
			      f"This generally means that the trade value does not match the holding value.")
			input("Press any key to continue, the program will exit after this and allow for secondary checks.")
			sys.exit("THERE HAS BEEN A PROBLEM WITH SELL ALL'S NOT BEING NOTED IN NOTES AND/OR VALUATION"
			         "PLEASE RESOLVE THIS ISSUE BEFORE CONTINUING.")
	else:
		return "Rebalancing"


def currency_finder(trade):
	"""
	The finds the currency that a trade is given in. If there is no currency found, assume GBP.
	:param trade: Cell containing trade.
	:return: Returns currency.
	"""
	if trade.offset(0, -4).value is not None:
		return trade.offset(0, -4).value
	else:
		return "GBP"


def _notes_value(trade):
	"""
	Under development:
	Takes trade value, returns notes from context based locations.
	:param trade: Cell containing trade
	:return: Returns contents for notes column.
	"""
	notes = trade.offset(0, 5).value
	if notes is not None:
		# This is another quick sanity check. If The column contains 'sell' indicating that this is a sell all
		# But the trade is a buy, then it will alert the user to this, and give the user the
		if "broker" in notes.lower() or "etf" in notes.lower():
			return "Broker Trade"
		if notes.lower() == "cis":
			return "Rebalancing"
		if "sell all" in notes.lower():
			return "Sell All"
		return notes
	if sell_all_reader(trade).lower() == "sell all":
		return "Sell All"
	else:
		return 'Rebalancing'


# This generates the list of lists to be written to the csv document.
def automated_trading_array_generator(trade_array):
	""" Takes an array of all trade locations. Returns an array that contains all trading information. """
	trades = []
	for idx, a in enumerate(trade_array, start=1):
		trade_values = trade_array_generator(idx, a)
		trades.append(trade_values)
	else:
		return trades


def main_writer(df, headers):
	global NON_BROKER_FILENAME
	print(F"In here. {df}")
	NON_BROKER_FILENAME = str('Unchecked_' + FUND_MANAGER + '_' + date.today().strftime("%Y%m%d") + '.csv')
	if exists(NON_BROKER_FILENAME):
		headers = False
		mode = 'a'
	else:
		mode = 'w'
	df.to_csv(NON_BROKER_FILENAME, index=False, mode=mode, header=headers)
	return NON_BROKER_FILENAME
