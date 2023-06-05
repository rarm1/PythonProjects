# TODO: Remove single use values
# TODO: Refactor to be OO
import logging
import sys
from datetime import date
from os.path import exists

import openpyxl as xl
import pandas as pd

from assets import constants

# Variable initialisation
global MODIFIED_FILENAME, WRITER_SHEET, READER_SHEET, READER_DOC, NON_BROKER_FILENAME, BROKER_FILENAME, LOCAL_EXEC, \
	FUND_MANAGER
TRADES = 0
# These are used so that if the location changes, it can be changed here, without refactoring the entire project.
SCHEME_LIST_LOCATION = 'Resources/scheme_list_desig.xlsx'
TRADE_ID_FILE = 'Resources/trade_identifier.xlsx'
TRADE_ID_SHEET = 'trade_identifier'

# The scheme loopup is written here. This is for ease of updating.
# Again, if multiple people are using this program, then
# ensuring every instance of this document is updated will be more error-prone.
SCHEME_LOOKUP_DOC = xl.load_workbook(SCHEME_LIST_LOCATION)
SCHEME_SHEET = SCHEME_LOOKUP_DOC['Source']
ISIN_TO_TICKER_PATH = 'X:/Fund Management/Fund Management Team Files/Richard/ISIN_to_Ticker.xlsx'
ISIN_TO_TICKER_DOC = xl.load_workbook(ISIN_TO_TICKER_PATH, data_only=True)
ISIN_TO_TICKER_SHEET = ISIN_TO_TICKER_DOC['Sheet1']
FOLDER_DATE = date.today().strftime("%Y%m%d")
FUND_MANAGERS = {
	"Liz": "ELIZABETH",
	"Dmitry": "DMITRY",
	"Richard Cole": "FMONEY",
	"Wayne": "OPES",
	"IBOSS": "IBOSS",
	"Sentinel": "SENTINEL"
}
FUND_MANAGER_VALUES = list(FUND_MANAGERS.values())
HEADERS_GENERATOR = ['FMDealingCode', 'TradeIdentifier', 'TradeFileDate', 'SchemeIdentifier', 'SecurityName',
                     'SecurityIdentifier', 'NewSecurity', 'Direction', 'DealType', 'Value',
                     'SettlementCurrency', 'PreviouslyReported', 'Notes',
                     'Scheme from Rebalancer (DELETE AFTER VERIFICATION)',
                     'Scheme from PID lookup sheet (DELETE AFTER VERIFICATION)']
BROKER_HEADERS_GENERATOR = ['Thomas Grant Account Number', 'Ticker', 'Fund Designation Number', 'FMDealingCode',
                            'TradeIdentifier',
                            'TradeFileDate', 'SchemeIdentifier', 'SecurityName',
                            'SecurityIdentifier', 'NewSecurity', 'Direction', 'DealType', 'Value',
                            'Quantity of Units To Trade', 'Approx Price Per Unit last VP', 'Notes',
                            'Scheme from Rebalancer (DELETE AFTER VERIFICATION)',
                            'Scheme from PID lookup sheet (DELETE AFTER VERIFICATION)']


def single_trade_writer(fund, trade_amount):
	"""
	This accepts a trade and appends it to the rebalancer sheet.
	:param fund: Cell
	:param trade_amount: Int
	:return: Void
	"""
	global TRADES
	row = fund.row
	column = fund.column + 2
	reader_cell = READER_SHEET.cell(row=row, column=column)
	if trade_amount < 0 and (
			(fund.offset(0, -6).value - abs(trade_amount))
			/ constants.total_portfolio_value().value < 0.0005):
		reader_cell.offset(0, 3).value = "Sell All"
		trade_amount = -fund.offset(0, -6).value
	else:
		trade_amount = round(trade_amount, -3)
	reader_cell.value = trade_amount
	constants.cash_figure().value -= trade_amount
	TRADES += 1


def final_trade_writer():
	"""

	"""
	for fund in constants.all_funds():
		if fund.offset(0, 2).value is not None:
			row = fund.row
			column = fund.column
			WRITER_SHEET.cell(row=row, column=column + 2).value = fund.offset(0, 2).value
		if fund.offset(0, 5).value is not None:
			row = fund.row
			column = fund.column
			WRITER_SHEET.cell(row=row, column=column + 5).value = fund.offset(0, 5).value


def trade_finder():
	"""

	:return:
	"""
	all_trades = []
	for fund in constants.all_funds():
		if fund.offset(0, 2).value is not None:
			all_trades.append(fund)
	else:
		return all_trades


TODAY = date.today()
REQ_DATE = TODAY.strftime("%d/%m/%Y")


# This generator creates an array.
# Moving this function allows for more processing without the function getting overly complex.
def unit_price(trade):
	"""

	:param trade:
	:return:
	"""
	return abs(trade.offset(0, -6).value) / abs(trade.offset(0, -5).value)


def units_to_trade(price_per_unit, trade):
	"""

	:param price_per_unit:
	:param trade:
	:return:
	"""
	return abs(trade.offset(0, 2).value) / abs(price_per_unit)


def trade_array_generator(idx, trade):
	"""
	:param idx: This is an integer value.
	:param trade: This is a cell in which there is a trade.
	:return: Outputs a single completed trade array.
	"""
	global FUND_MANAGER, READER_SHEET
	sec_name = (trade.offset(0, -7)).value
	isin = trade.offset(0, -9).value
	deal_type = unit_cash(trade)
	trade_value = value_of_trade(deal_type, trade)
	trade_id = get_trade_value() + idx
	notes = notes_value(trade)
	if "broker" in notes.lower():
		# This is a broker trade, so it needs to be written to another sheet.
		price_per_unit = unit_price(trade)
		total_units_to_trade = units_to_trade(price_per_unit, trade)
		total_units = trade.offset(0, -5).value
		if abs(total_units - total_units_to_trade) < 0.01 * total_units:
			total_units_to_trade = total_units
		return True, ['#{0}'.format(str(f'{tg_account_number_lookup()}')),
		              ticker_lookup(isin), READER_SHEET.cell(4, 2).value, FUND_MANAGER,
		              trade_id, REQ_DATE,
		              scheme_lookup().value, sec_name,
		              isin, new_buy_finder(trade), buy_or_sell(trade), deal_type, abs(trade.offset(0, 2).value),
		              total_units_to_trade, price_per_unit, notes, READER_SHEET.cell(2, 2).value,
		              scheme_lookup().offset(0, -2).value]
	# The only entry in this array that is 'hardcoded' is the previousreported entry. This is because there is no
	# way of checking this, currently as discussed with Paul. It is also highly unlikely to be hit.
	# So return on time invested is low. #
	return False, [FUND_MANAGER, trade_id, REQ_DATE, scheme_lookup().value, sec_name,
	               isin, new_buy_finder(trade), buy_or_sell(trade), deal_type, trade_value,
	               currency_finder(trade), "N", notes, READER_SHEET.cell(2, 2).value,
	               scheme_lookup().offset(0, -2).value]


def data_frame_generator(trades):
	"""
	:param trades: accepts an array of trades.
	:return: Returns a formatted list of lists that gives all required information
	"""
	to_return = []
	for trade in trades:
		fund_name = trade.offset(0, -7).value
		isin = trade.offset(0, -9).value
		new_buy = new_buy_finder(trade)
		buy_or_sell_var = buy_or_sell(trade)
		unit_or_cash_var = trade_type(trade)
		currency = currency_finder(trade)
		to_return.append(["Fund Manager Placeholder", "Trade ID Placeholder", "Todays Date", "Scheme Lookup", fund_name,
		                  isin, new_buy, buy_or_sell_var, unit_or_cash_var, trade.offset(0, 2).value, currency, "N",
		                  "Notes Placeholder", "Rebalancer Scheme Placeholder", "Rebalancer Lookup Placeholder"])
	else:
		return to_return


# This requests the name of the fund manager from the user and uses this to write to the csv.
def fundmanager_selection():
	"""

	:return:
	"""
	global FUND_MANAGER
	# The use of enumerate here allows for easier visualisation of the array for the user.
	# without using this function the list would start at '0' which is much less natural for most potential users. #
	for idx, x in enumerate(FUND_MANAGERS, start=1):
		print(idx, x)
	while True:
		desired_index = input("Input the number of the fund manager to select: ")
		try:
			desired_index = int(desired_index)
			if len(FUND_MANAGERS) + 1 > desired_index > 0:
				break
		except ValueError:
			print("NaN")
	print("\n\n")
	FUND_MANAGER = FUND_MANAGER_VALUES[desired_index - 1]
	return FUND_MANAGER_VALUES[desired_index - 1]


# FUND_MANAGER = fundmanager_selection()


# This function sets fund managers for third parties, or instructed trades.
# e.g. Clarion may instruct trades, but Margetts
# We do the rest of the paperwork. As such, Elizabeth is going to be the fund manager.
# Does this need additional oversight?
# Are IBOSS / Clarion likely to be the FMs? #
def get_fund_manager():
	"""

	:return:
	"""
	return FUND_MANAGER


# This finds the scheme loopup reference document. This document, again, is located in a central place, such that it can
# be updated. Provided the layout, and name of the file is the same, then it can be changed without modifying any code.
def scheme_lookup():
	"""

	:return:
	"""
	rebalancer_scheme = READER_SHEET.cell(4, 2)
	for x in range(1, SCHEME_SHEET.max_row + 1):
		scheme_pid_lookup = (SCHEME_SHEET.cell(x, 2))
		if str(scheme_pid_lookup.value) == str(rebalancer_scheme.value):
			return scheme_pid_lookup.offset(0, 1)
	return f"Scheme PID {rebalancer_scheme.value} not listed in {SCHEME_LIST_LOCATION}"


def ticker_lookup(isin):
	"""

	:param isin:
	:return:
	"""
	for i in range(1, ISIN_TO_TICKER_SHEET.max_row + 1):
		isin_in_lookup = ISIN_TO_TICKER_SHEET.cell(i, 1)
		if isin_in_lookup.value is not None:
			if isin.lower() in isin_in_lookup.value.lower():
				return ISIN_TO_TICKER_SHEET.cell(i, 2).value
	else:
		return "ISIN TO BE ADDED"


def tg_account_number_lookup():
	"""

	:return:
	"""
	rebalancer_scheme = READER_SHEET.cell(2, 2)
	for x in range(1, TG_ACCOUNT_NUMBERS.max_row + 1):
		if TG_ACCOUNT_NUMBERS.cell(x, 1).value is not None:
			if TG_ACCOUNT_NUMBERS.cell(x, 1).value.lower() in rebalancer_scheme.value.lower():
				return TG_ACCOUNT_NUMBERS.cell(x, 2).value


def new_buy_finder(trade):
	"""
	Determines whether the fund is already held or whether it is a new fund in the portfolio
	:param trade: Cell containing trade
	:return: Y or N based on new holding or not.
	"""
	if trade.offset(0, -2).value == 0:
		return "Y"
	else:
		return "N"


# Define buy or sell. Is the trade greater or larger than 0.
def buy_or_sell(trade):
	"""
	Takes a trade and returns whether it is a buy or a sell.
	:param trade: Cell containing trade
	:return: Direction of trade.
	"""
	if trade.offset(0, 2).value > 0:
		return "BUY"
	else:
		return "SELL"


# This is the trade identifier that is generated in the constants file.
# The reason that this file is centralised is that it reduces duplicate trade errors.
def get_trade_value():
	"""

	:return:
	"""
	trade_counter_doc = xl.load_workbook(TRADE_ID_FILE)
	trade_sheet = trade_counter_doc[TRADE_ID_SHEET]
	return trade_sheet.cell(1, 1).value


# This returns the value of the trade. The reason this function exists instead of reading a set value is that if
# all the fund is being sold, then this function must return the amount of units held in the fund.
def value_of_trade(deal_type, trade):
	"""
	Ensure that the correct trade value is returned. This changes based on whether a trade is currency or units.
	:param deal_type: Units or cash
	:param trade: Cell containing trade
	:return: Returns the quantity to be traded.
	"""
	if deal_type == "UNIT":
		return trade.offset(0, -5).value
	else:
		return abs(trade.offset(0, 2).value)


def unit_cash(trade):
	"""
	Generate whether this function is a cash deal or a unit deal.
	:param trade: Cell containing trade.
	:return: Returns units or cash.
	"""
	if abs(round(trade.offset(0, 3).value, 4)) == 0:
		return "UNIT"
	else:
		return "CASH"


def sell_all_reader(trade):
	"""
	Determines whether it is a sell all or a normal rebalancing trade.
	:param trade: Cell containing trade
	:return: Returns either sell all or rebalancing
	"""
	if abs(round(trade.offset(0, 3).value, 4)) == 0:
		return 'Sell All'
	else:
		return "Rebalancing"


def trade_type(trade):
	"""
	Determines whether a trade returns units or cash.
	:param trade: The cell that contains a trade.
	:return: Either cash or units.
	"""
	if abs(abs(trade.offset(0, 2).value) - trade.offset(0, -6).value) > constants.total_portfolio_value().value:
		return "UNIT"
	else:
		return "CASH"


# This functions takes advantage the lastest rebalancer template, which provides currency and units for each fund.
# The 'default' setting for this function is GBP, as that is the vast majority of funds that Margetts are invested in.
def currency_finder(trade):
	"""
	The finds the currency that a trade is given in. If there is no currency found, assume GBP.
	:param trade: Cell containing trade.
	:return: Returns currency.
	"""
	if trade.offset(0, -4).value is not None:
		return trade.offset(0, -4).value
	else:
		return "GBP"


"""This generates 'notes' for the CSV. This function can be expanded to allow for Dmitry to instruct trades, etc.
This may however require an intermediary step to determine whether Dmitry is instructing trades, or acting as
an instructee.
I have currently considered as a workaround for this, to request, at Wednesday's meeting for rebalancing whether we
may use the column to the right of 'sector' column Q for a 'reason' column, if you will. So if the trade is instructed
or critical, then place notes in that column. However, if there is no contents in the column, then assume rebalancing
I think the value in this is that it allows for a variety of instructions to be passed in this column that the
Dealing team can expand at their leisure. As much, or as little detail can be added to this column. """


def notes_value(trade):
	"""
	Under development:
	Takes trade value, returns notes from context based locations.
	:param trade: Cell containing trade
	:return: Returns contents for notes column.
	"""
	notes = trade.offset(0, 5).value
	if notes is not None:
		# This is another quick sanity check. If The column contains 'sell' indicating that this is a sell all
		# But the trade is a buy, then it will alert the user to this, and give the user the
		if "broker" in notes.lower():
			return "Broker Trade"
		if notes.lower() == "cis":
			return "Rebalancing"
		if 'sell' in notes.lower() and buy_or_sell(trade) == "BUY":
			print(f"Are you sure the notes are correct? It looks like this trade is a buy. "
				  f"{trade.offset(0, -7).value}\n")
			if input("Do you want to continue? Y/N: ").lower() == "y":
				return notes
			else:
				sys.exit("Closing the program, come back soon!")
		else:
			return notes
	if sell_all_reader(trade).lower() == "sell all":
		return "Sell All"
	if FUND_MANAGER == "IBOSS":
		"""For example, we could ask in here, whether this is a trade that IBOSS is instructing, or whether it is for
		Rebalancing. As this happens once a month, approximately, according to an office poll, this functionality has
		not been added. I think for the inconvinience factor this would add to instructing a dozen trades, that
		it would be easier to open the completed file in Excel and modify that values manually. #
		notes = input('Is this an instructed trade or rebalancing?')
		if notes == 'instructed':
			notes = 'IBOSS Instructed Trade'
		else:
			notes = 'Rebalancing Trade'"""
		return "IBOSS instructed trade"
	elif FUND_MANAGER == 'CLARION':
		return 'Clarion instructed trade'
	else:
		return 'Rebalancing'


def trade_tab_generator(trades):
	"""
	Generates the 'trades' tab on rebalancer documents. So suggested trades can be compared against finished trades.
	:param trades: Completed array of trades.
	:return: No return statement.
	"""
	headers = HEADERS_GENERATOR
	df = pd.DataFrame(trades, columns=headers)
	with pd.ExcelWriter(MODIFIED_FILENAME, engine="openpyxl", mode='a', if_sheet_exists='replace') as writer_:
		pd.DataFrame(df.to_excel(writer_, index=False, sheet_name="Trades"))


# This function writes the finished array to a csv that consists of the fund_managers name, and the date.
def main_writer(trades):
	"""
	Takes in an array of trades, writes these trades to a .csv file.
	:param trades: Array of all trades.
	:return: No return statement.
	"""
	global NON_BROKER_FILENAME, HEADERS_GENERATOR
	# Do not create empty files.
	if len(trades) == 0:
		return
	elif not len(trades) == 0:
		# This creates a dateaframe from the list of lists generated.
		df = pd.DataFrame(trades)
		req_date: str = date.today().strftime("%d-%m-%Y")
		# Use the variables defined above to generate a csv document
		# using functions that're built into the pandas repository
		if FUND_MANAGER == "CLARION":
			NON_BROKER_FILENAME = 'Unchecked_' + "ELIZABETH_" + req_date + '.csv'
		else:
			NON_BROKER_FILENAME = 'Unchecked_' + FUND_MANAGER + '_' + req_date + '.csv'
		# If this file exists, then there should not be any additional headers, and it should append, rather than
		# overwrite.
		if exists(NON_BROKER_FILENAME):
			headers = False
			mode = 'a'
		# If the file doesn't exist then use the header array and use the 'writer' file.
		else:
			headers = HEADERS_GENERATOR
			mode = 'w'
		try:
			df.to_csv(NON_BROKER_FILENAME, index=False, mode=mode, header=headers)
		except PermissionError as e:
			logging.error("Error occured" + str(e))
			NON_BROKER_FILENAME = NON_BROKER_FILENAME[:-4] + '_V2.csv'
			df.to_csv(NON_BROKER_FILENAME, index=False, mode=mode, header=headers)
		return NON_BROKER_FILENAME


# This function writes the finished array to a csv that consists of the fund_managers name, and the date.
# def broker_trade_writer(trades):
# 	"""
# 	Takes in an array of trades, writes these trades to a .csv file.
# 	:param trades: Array of all trades.
# 	:return: No return statement.
# 	"""
# 	global BROKER_FILENAME
# 	# Do not create empty files.
# 	if len(trades) == 0:
# 		return
# 	elif not len(trades) == 0:
# 		# This creates a dateaframe from the list of lists generated.
# 		df = pd.DataFrame(trades)
# 		req_date: str = date.today().strftime("%d-%m-%Y")
# 		# Use the variables defined above to generate a csv document using
# 		# functions that're built into the pandas repository
# 		if FUND_MANAGER == "CLARION":
# 			BROKER_FILENAME = "Unchecked_ELIZABETH_" + req_date + '_broker_trades.csv'
# 		else:
# 			BROKER_FILENAME = 'Unchecked_' + FUND_MANAGER + '_' + req_date + '_broker_trades.csv'
# 		# If this file exists, then there should not be any additional headers, and it should append, rather than
# 		# overwrite.
# 		if exists(BROKER_FILENAME):
# 			headers = False
# 			mode = 'a'
# 		# If the file doesn't exist then use the header array and use the 'writer' file.
# 		else:
# 			headers = BROKER_HEADERS_GENERATOR
# 			mode = 'w'
# 		df.to_csv(BROKER_FILENAME, index=False, mode=mode, header=headers)
# 		return BROKER_FILENAME

#
# # This generates the list of lists to be written to the csv document.
# def automated_trading_array_generator(trade_array):
# 	""" Takes an array of all trade locations. Returns an array that contains all trading information. """
# 	non_broker_trades, broker_trades = [], []
# 	trade_id = 1
# 	constants.READER_SHEET = READER_SHEET
# 	for idx, a in enumerate(trade_array, start=1):
# 		trade_id = get_trade_value() + idx
# 		broker_trade, trial_return = trade_array_generator(idx, a)
# 		if broker_trade:
# 			broker_trades.append(trial_return)
# 		else:
# 			non_broker_trades.append(trial_return)
#
# 	else:
# 		if not LOCAL_EXEC and len(trade_array) != 0:
# 			write_trade_value(trade_id)
# 		return non_broker_trades, broker_trades


# This updates the trade number after ever rebalancer that has been read. This adds a stable point in which to rewrite.
# And doesn't give the user a chance to write trades, and then escape the program without updating the trade quantity.
# def write_trade_value(trade):
# 	"""
#
# 	:param trade:
# 	"""
# 	trade_counter_doc = xl.load_workbook(TRADE_ID_FILE)
# 	trade_sheet = trade_counter_doc[TRADE_ID_SHEET]
# 	trade_sheet.cell(1, 1).value = trade
# 	trade_counter_doc.save(TRADE_ID_FILE)
