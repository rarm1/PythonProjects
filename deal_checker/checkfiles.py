"""
Good afternoon! So, apologies for the way this is written. I did it in about 30 minutes, and it works!
"""
import openpyxl as xl
from Resources import constants

'''
Master column 7 is Prudence (g) 
Master column 11 is Navigator (k)
Master column 15 is Meridian (o)
Master column 19 is Explorer (s)
'''
# These filenames could be changed, or files could be saved with alternative filenames.
funds = [[7, "Clarion Prudence 08 07 2022.xlsx"], [11, "Clarion Navigator 08 07 2022.xlsx"],
         [15, "Clarion Meridian 08 07 2022.xlsx"], [19, "Clarion Explorer 08 07 2022.xlsx"]]

# List these files so the user can select a file.
for idx, fund in enumerate(funds):
    print(idx, fund[1])

# No sanitation, assume that the user isn't going to try and trip you up. If they are, then they suffer more!
desired_file = int(input("Please entire file number:"))
desired_file = funds[desired_file]
# Load document with openpyxl. Data_only reads only data from excel rather than formulas.
master_trade_doc = xl.load_workbook("Model Stages_2.xlsx", data_only=True)
# This sheetname can be changed if trades are on another sheet in Excel. As seen on "Model Stages_2". Our trades are
#   On a sheet called "Day 1" Mirror this format.
master_trade_sheet = master_trade_doc['Day 1']
rebalancer_doc = xl.load_workbook(desired_file[1], data_only=True)
rebalancer_sheet = rebalancer_doc['Rebalancer']
# References a sheet in the 'Resources' folder. This was the backbone of a lot of work I did, so, I reused here.
constants.READER_SHEET = rebalancer_sheet
# Arbitrary counters to give an idea how close we are to perfect trades.
incorrect_trades = 0
correct_trades = 0
# The magic happens.
# The all funds function will retreive every fund from the rebalancer given above. It is not important how it finds each
#   fund.
for trade in constants.all_funds():
    # This loops every trade on the sheet.
    # Using the given row, I find the ISIN by referencing the third column.
    isin_of_trade = rebalancer_sheet.cell(row=trade.row, column=3)
    # If the value is not none, then it should be a number. This could cause problems if it split out and read rows
    #   That it shouldn't & so wrapping in a try except could be helpful.
    if trade.offset(0, 3).value is not None:
        trade_quantity = round(trade.offset(0, 3).value, 4)
    else:
        # This basically ignores the trade if it's not here, or will raise a large enough discrepancy to get attention.
        trade_quantity = 0
    # In a rather in-eligant fashion, this loops through the model spreadsheet for every trade on the rebalancer.
    #   We're not running this code thousands of times, the sheets aren't huge, I had to do it quickly, nothing is order.
    #   Pick an excuse and I'm happy with it. Also, PEP 8 told me to keep it simple, stupid. So I am.
    for row_num in range(1, master_trade_sheet.max_row):
        master_trade_isin = master_trade_sheet.cell(row=row_num, column=1)
        # Same logic here.
        if master_trade_sheet.cell(row=row_num, column=desired_file[0] + 1).value is not None:
            try:
                # I've used a try except here because I expect errors to be thrown. Above I used lots of logic to determine
                #   Where the funds start and end. I do not use the same logic on this model sheet as that'd take too long.
                #   Therefore, I need to be aware that text will be passed through this, and it must be caught.
                #   It is not an error, hence why I just continue to the next iteration.
                master_trade_quantity = round(
                    master_trade_sheet.cell(row=row_num, column=desired_file[0] + 1).value, 4)
            except TypeError as e:
                # print(e)
                continue
        else:
            master_trade_quantity = 0
        # Seeing as ISINs are unique, this means 'a bingo'. So we should compare the trades from these rows.
        if isin_of_trade.value == master_trade_isin.value:
            # If they are not IDENTICAL which is kind of gross. Maybe doing an equal or plus or minus 20 basis points.
            #   But again, speed.
            if not trade_quantity == master_trade_quantity:
                # This is all within a try catch, which arguably isn't necessary.
                print(isin_of_trade.offset(0, 2).value)
                try:
                    # This is, I suppose is kind of the point of the program. It very simply, prints out the targets
                    #   gathered from the model deals sheet, and the POST TRADE percent from the rebalancer.
                    #   This allows for easy visualisation of the differences between the two figures.
                    print(f"Post trade pct from the rebalancer is {round(trade_quantity * 100, 4)}%.\n"
                          f"The target from the model trade quantity is {round(master_trade_quantity * 100, 4)}%\n"
                          f"Difference between these is {abs(round((trade_quantity - master_trade_quantity) * 100, 4))}%\n"
                          f"On the model sheet, this will be on row {master_trade_isin.row}.\n"
                          f"On the rebalancer this will be on {isin_of_trade.row}\n")
                    incorrect_trades += 1
                except TypeError as e:
                    print(f"Rebalancer trade is {trade_quantity}. Model trade quantity is {master_trade_quantity}")
                    print(
                        f"On the model sheet, this will be on row {master_trade_isin.row}. On the rebalancer this will be"
                        f" on {isin_of_trade.row}\n")
            else:
                correct_trades += 1

print(f"For a total of {correct_trades} correct trades & {incorrect_trades} incorrect trades")
